from __future__ import annotations

import json
import zipfile
from pathlib import Path

import pytest

from workers_projects_runtime.deliverables import candidate_artifact_paths, is_user_deliverable_relative_path
from workers_projects_runtime.runtime_identity import derive_legacy_backend_label
import workers_projects_runtime.run_evidence as run_evidence
from workers_projects_runtime.run_evidence import (
    _text_artifact_payload,
    build_constraint_ledger,
    build_run_evidence,
    check_content_hygiene,
    summarize_run_evidence_result,
    write_constraint_ledger,
    write_run_evidence,
)


def test_constraint_ledger_extracts_generic_source_date_and_flag_rules(tmp_path):
    instruction = (
        "Research the target market using sources from January 2024 through May 2026 only.\n"
        "Include seed firms Alpha Capital and Beta Partners.\n"
        "Do not exclude possible platform partners; flag them instead.\n"
        "Deliver an XLSX and PDF report."
    )

    ledger = build_constraint_ledger(
        instruction=instruction,
        worker={"worker_id": "wrk_ledger", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_ledger",
    )

    assert ledger["run_id"] == "run_ledger"
    assert ledger["worker"]["profile"] == "codex-cli"
    assert any("January 2024 through May 2026 only" in item for item in ledger["constraints"]["date"])
    assert any("sources from" in item for item in ledger["constraints"]["source"])
    assert any("Do not exclude" in item for item in ledger["constraints"]["exclusion_or_flag"])
    assert any("XLSX and PDF" in item for item in ledger["outputs"]["required"])
    assert ledger["do_not_widen_or_soften"] is True

    latest = write_constraint_ledger(tmp_path, ledger, "run_ledger")
    assert latest == tmp_path / "glasshive-run" / "constraint-ledger.json"
    assert json.loads(latest.read_text())["run_id"] == "run_ledger"
    assert (tmp_path / "glasshive-run" / "runs" / "run_ledger" / "constraint-ledger.json").exists()


def test_constraint_ledger_does_not_treat_uploaded_as_date_constraint():
    ledger = build_constraint_ledger(
        instruction=(
            "Inspect the uploaded PDF directly inside the workspace under uploads/<filename>. "
            "Create a concise artifact named upload-fallback-ui-smoke.txt that states whether the first "
            "bytes are the PDF signature %PDF. Keep working until the user's request is satisfied."
        ),
        worker={
            "worker_id": "wrk_upload_ledger",
            "profile": "codex-cli",
            "execution_mode": "host",
        },
        run_id="run_upload_ledger",
    )

    assert ledger["constraints"]["date"] == []
    assert ledger["outputs"]["format_expectations"] == ["txt"]
    assert all("<filename>" not in seed for seed in ledger["seed_entities_or_files"])


def test_run_evidence_accepts_uploaded_pdf_input_with_text_artifact(tmp_path):
    uploads = tmp_path / "uploads"
    uploads.mkdir()
    (uploads / "input-smoke.pdf").write_bytes(b"%PDF-1.4\n% synthetic\n")
    (tmp_path / "upload-fallback-ui-smoke.txt").write_text(
        "uploaded file name: input-smoke.pdf\n"
        "byte size: 21\n"
        "first bytes are %PDF: yes\n"
        "workspace path: uploads/input-smoke.pdf\n"
    )
    worker = {
        "worker_id": "wrk_upload_evidence",
        "profile": "codex-cli",
        "execution_mode": "host",
        "bootstrap_bundle_json": json.dumps({"files": [{"path": "uploads/input-smoke.pdf"}]}),
    }
    ledger = build_constraint_ledger(
        instruction=(
            "Inspect the uploaded PDF directly inside the workspace under uploads/<filename>. "
            "Create a concise artifact named upload-fallback-ui-smoke.txt that states whether the first "
            "bytes are the PDF signature %PDF."
        ),
        worker=worker,
        run_id="run_upload_evidence",
    )

    evidence = build_run_evidence(
        worker=worker,
        run_id="run_upload_evidence",
        runtime_name="codex-cli",
        model="gpt-test",
        command=["codex", "exec", "Inspect upload."],
        env={},
        workspace_dir=tmp_path,
        stdout_text="FINAL REPORT:\nCreated upload-fallback-ui-smoke.txt",
        stderr_text="",
        output_text="FINAL REPORT:\nCreated upload-fallback-ui-smoke.txt",
        error_text="",
        exit_code=0,
        timeout_seconds=300,
        stop_reason="process_exit",
        constraint_ledger=ledger,
    )

    assert evidence["completion_compliance"]["required_artifact_types"] == ["txt"]
    assert evidence["completion_compliance"]["seed_entity_coverage"]["status"] == "pass"
    assert evidence["evidence_result"]["status"] == "pass"


def test_run_evidence_does_not_require_original_upload_format_when_html_requested(tmp_path):
    uploads = tmp_path / "uploads"
    uploads.mkdir()
    (uploads / "input-smoke.pdf").write_bytes(b"%PDF-1.4\n% synthetic\n")
    (tmp_path / "upload-smoke.html").write_text(
        "FINAL REPORT:\n"
        "filename: input-smoke.pdf\n"
        "byte size: 21\n"
        "first bytes: %PDF-\n"
        "marker: GH_UPLOAD_SMOKE_A\n"
    )
    worker = {
        "worker_id": "wrk_upload_html_evidence",
        "profile": "codex-cli",
        "execution_mode": "host",
        "bootstrap_bundle_json": json.dumps({"files": [{"path": "uploads/input-smoke.pdf"}]}),
    }
    ledger = build_constraint_ledger(
        instruction=(
            "Use GlassHive on the attached PDF. Verify the actual uploaded PDF bytes are present "
            "in the worker workspace uploads folder. Create an HTML artifact named upload-smoke.html "
            "containing filename, byte size, %PDF first-byte check, exact marker GH_UPLOAD_SMOKE_A, "
            "and one visual redesign note. Return the artifact link. If the original PDF bytes are "
            "missing, say that as the blocker. "
            "Attached file reference: input-smoke.pdf. Known marker embedded in the PDF: GH_UPLOAD_SMOKE_A. "
            "Please create the requested HTML artifact only if the bytes are present and verifiable. "
            "Attached/uploaded file reference: input-smoke.pdf. The known embedded marker is GH_UPLOAD_SMOKE_A. "
            "The uploaded file content visible in chat context includes: 'Synthetic PDF proves byte-level "
            "upload materialization. Expected worker behavior: read this PDF from uploads/ and create an artifact.' "
            "Preserve the user's exact constraint: create the requested HTML artifact only if the bytes are "
            "present and verifiable. Keep working until the user's request is satisfied."
        ),
        worker=worker,
        run_id="run_upload_html_evidence",
    )

    assert ledger["constraints"]["date"] == []
    assert ledger["outputs"]["format_expectations"] == ["html"]
    assert "Preserve the user" not in run_evidence._seed_terms_from_ledger(ledger)

    evidence = build_run_evidence(
        worker=worker,
        run_id="run_upload_html_evidence",
        runtime_name="codex-cli",
        model="gpt-test",
        command=["codex", "exec", "Inspect upload."],
        env={},
        workspace_dir=tmp_path,
        stdout_text="FINAL REPORT:\nCreated upload-smoke.html",
        stderr_text="",
        output_text="FINAL REPORT:\nCreated upload-smoke.html",
        error_text="",
        exit_code=0,
        timeout_seconds=300,
        stop_reason="process_exit",
        constraint_ledger=ledger,
    )

    completion = evidence["completion_compliance"]
    assert completion["required_artifact_types"] == ["html"]
    assert "pdf" not in completion["required_artifact_types"]
    assert completion["missing_required_artifact_types"] == []
    assert completion["status"] == "pass"
    assert evidence["evidence_result"]["status"] == "pass"


def test_run_evidence_fails_when_verified_artifact_lacks_final_marker(tmp_path):
    uploads = tmp_path / "uploads"
    uploads.mkdir()
    (uploads / "input-smoke.pdf").write_bytes(b"%PDF-1.4\n% synthetic\n")
    (tmp_path / "upload-smoke.html").write_text(
        "filename: input-smoke.pdf\n"
        "byte size: 21\n"
        "first bytes: %PDF-\n"
        "marker: GH_UPLOAD_SMOKE_A\n"
    )
    worker = {
        "worker_id": "wrk_upload_html_no_final",
        "profile": "codex-cli",
        "execution_mode": "host",
        "bootstrap_bundle_json": json.dumps({"files": [{"path": "uploads/input-smoke.pdf"}]}),
    }
    ledger = build_constraint_ledger(
        instruction=(
            "Use the attached PDF and create an HTML artifact named upload-smoke.html containing "
            "filename, byte size, %PDF first-byte check, and exact marker GH_UPLOAD_SMOKE_A. "
            "Attached/uploaded file reference: input-smoke.pdf. Expected worker behavior: read this "
            "PDF from uploads/ and create an artifact."
        ),
        worker=worker,
        run_id="run_upload_html_no_final",
    )

    evidence = build_run_evidence(
        worker=worker,
        run_id="run_upload_html_no_final",
        runtime_name="codex-cli",
        model="gpt-test",
        command=["codex", "exec", "Inspect upload."],
        env={},
        workspace_dir=tmp_path,
        stdout_text="Created upload-smoke.html",
        stderr_text="",
        output_text="Created upload-smoke.html",
        error_text="",
        exit_code=0,
        timeout_seconds=300,
        stop_reason="process_exit",
        constraint_ledger=ledger,
    )

    completion = evidence["completion_compliance"]
    assert completion["required_artifact_types"] == ["html"]
    assert completion["missing_required_artifact_types"] == []
    assert completion["status"] == "fail"
    assert evidence["evidence_result"]["status"] == "fail"
    assert any(
        reason["reason"] == "final report marker missing"
        for reason in evidence["evidence_result"]["failure_reasons"]
    )


@pytest.mark.parametrize(
    ("instruction", "expected"),
    [
        ("Create an HTML summary of the attached PDF.", ["html"]),
        ("Build an xlsx workbook from the uploaded csv source.", ["xlsx"]),
        ("Write a docx summary referencing the original pdf.", ["docx"]),
        ("Save the result as dashboard.html using the attached pdf.", ["html"]),
        ("Create a PDF report from the attached spreadsheet.", ["pdf"]),
    ],
)
def test_constraint_ledger_keeps_requested_output_formats_without_requiring_input_formats(instruction, expected):
    ledger = build_constraint_ledger(
        instruction=instruction,
        worker={"worker_id": "wrk_output_formats", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_output_formats",
    )

    assert ledger["outputs"]["format_expectations"] == expected


def test_constraint_ledger_ignores_mime_and_pdf_bytes_host_upload_guidance():
    instruction = "\n".join(
        [
            (
                "Use GlassHive on the attached PDF. Verify the actual uploaded PDF bytes are present in "
                "the worker workspace uploads folder. Create an HTML artifact named upload-final-smoke.html "
                "containing filename, byte size, %PDF first-byte check, exact marker GH_UPLOAD_SMOKE_A, "
                "and one visual redesign note. Return the artifact link. If the original PDF bytes are "
                "missing, say that as the blocker."
            ),
            (
                "Attached file reference: upload-smoke.pdf (application/pdf). The worker should inspect the "
                "workspace uploads folder directly and verify actual bytes, not metadata only. If available, "
                "verify the file starts with %PDF and check for the exact marker string GH_UPLOAD_SMOKE_A in "
                "the PDF bytes/content. Then create the requested HTML artifact named upload-final-smoke.html "
                "and provide its artifact link. If the PDF bytes are missing or inaccessible, report that as "
                "the blocker rather than fabricating results."
            ),
        ]
    )

    ledger = build_constraint_ledger(
        instruction=instruction,
        worker={"worker_id": "wrk_host_upload_guidance", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_host_upload_guidance",
    )

    assert ledger["outputs"]["format_expectations"] == ["html"]


def test_run_evidence_recursively_validates_artifacts_and_flags_date_drift(tmp_path):
    output_dir = tmp_path / "output"
    output_dir.mkdir()
    (output_dir / "screen.csv").write_text("firm,notes\nAlpha,ok\n")
    (output_dir / "report.pdf").write_bytes(b"%PDF-1.7\n% synthetic\n")
    with zipfile.ZipFile(output_dir / "workbook.xlsx", "w") as archive:
        archive.writestr("[Content_Types].xml", "<Types></Types>")
        archive.writestr("xl/workbook.xml", "<workbook></workbook>")
    research = tmp_path / "research"
    research.mkdir()
    (research / "SPEC.md").write_text("Use sources dated Jan 2024 - June 2026 wherever possible.\n")
    (tmp_path / "glasshive-run").mkdir()
    (tmp_path / "glasshive-run" / "internal.txt").write_text("not a user artifact")

    ledger = build_constraint_ledger(
        instruction="Use sources from January 2024 through May 2026 only.",
        worker={"worker_id": "wrk_artifacts", "profile": "claude-code", "execution_mode": "host"},
        run_id="run_artifacts",
    )

    evidence = build_run_evidence(
        worker={"worker_id": "wrk_artifacts", "profile": "claude-code", "execution_mode": "host"},
        run_id="run_artifacts",
        runtime_name="claude-code",
        model="claude-opus-4-8",
        command=["claude", "-p", "--effort", "max"],
        env={"GLASSHIVE_RUN_ID": "run_artifacts", "SECRET_TOKEN": "private"},
        workspace_dir=tmp_path,
        stdout_text="FINAL REPORT:\ncreated files",
        stderr_text="debug image path /Users/example/private-workspace/tmp/preview.png",
        output_text="FINAL REPORT:\ncreated files",
        error_text="",
        exit_code=0,
        timeout_seconds=None,
        stop_reason="",
        constraint_ledger=ledger,
    )

    artifact_paths = {item["path"] for item in evidence["artifacts"]["items"]}
    assert "output/screen.csv" in artifact_paths
    assert "output/report.pdf" in artifact_paths
    assert "output/workbook.xlsx" in artifact_paths
    assert "glasshive-run/internal.txt" not in artifact_paths
    workbook = next(item for item in evidence["artifacts"]["items"] if item["path"] == "output/workbook.xlsx")
    assert workbook["document_validation"]["valid"] is True
    assert "SECRET_TOKEN" not in evidence["env_keys"]
    assert evidence["final_output"]["has_final_report"] is True
    assert evidence["constraint_compliance"]["status"] == "fail"
    assert any("June 2026" in issue["text"] for issue in evidence["constraint_compliance"]["issues"])
    assert evidence["evidence_result"]["status"] == "fail"
    assert any(item["reason"] == "constraint compliance failed" for item in evidence["evidence_result"]["failure_reasons"])

    latest = write_run_evidence(tmp_path, evidence, "run_artifacts")
    assert json.loads(latest.read_text())["run_id"] == "run_artifacts"
    assert (tmp_path / "glasshive-run" / "runs" / "run_artifacts" / "evidence.json").exists()


def _write_minimal_pdf(path: Path, title: str = "GlassHive test PDF") -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    content = f"BT /F1 18 Tf 72 720 Td ({title}) Tj ET".encode("utf-8")
    objects = [
        b"<< /Type /Catalog /Pages 2 0 R >>",
        b"<< /Type /Pages /Kids [3 0 R] /Count 1 >>",
        b"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] /Resources << /Font << /F1 4 0 R >> >> /Contents 5 0 R >>",
        b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>",
        b"<< /Length %d >>\nstream\n" % len(content) + content + b"\nendstream",
    ]
    body = b"%PDF-1.4\n"
    offsets = [0]
    for index, obj in enumerate(objects, start=1):
        offsets.append(len(body))
        body += f"{index} 0 obj\n".encode("ascii") + obj + b"\nendobj\n"
    xref_offset = len(body)
    body += f"xref\n0 {len(objects) + 1}\n0000000000 65535 f \n".encode("ascii")
    for offset in offsets[1:]:
        body += f"{offset:010d} 00000 n \n".encode("ascii")
    body += (
        f"trailer\n<< /Size {len(objects) + 1} /Root 1 0 R >>\n"
        f"startxref\n{xref_offset}\n%%EOF\n"
    ).encode("ascii")
    path.write_bytes(body)


def test_run_evidence_transcript_metadata_resolves_workspace_relative_paths(tmp_path):
    ledger = build_constraint_ledger(
        instruction="Deliver a short Markdown note.",
        worker={"worker_id": "wrk_transcript_meta", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_transcript_meta",
    )
    ledger_path = write_constraint_ledger(tmp_path, ledger, "run_transcript_meta")
    stdout_path = tmp_path / "stdout.log"
    stdout_path.write_text("FINAL REPORT:\nDone.\n")

    evidence = build_run_evidence(
        worker={"worker_id": "wrk_transcript_meta", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_transcript_meta",
        runtime_name="codex-cli",
        model="gpt-5.4",
        command=["codex", "exec"],
        env={},
        workspace_dir=tmp_path,
        stdout_text=stdout_path.read_text(),
        stderr_text="debug image path /Users/example/private-workspace/tmp/preview.png",
        output_text="FINAL REPORT:\nDone.",
        error_text="",
        exit_code=0,
        timeout_seconds=None,
        stop_reason="",
        constraint_ledger=ledger,
        transcript_paths={
            "constraint_ledger": ledger_path.relative_to(tmp_path).as_posix(),
            "stdout": str(stdout_path),
        },
    )

    metadata = evidence["transcript"]["metadata"]
    assert metadata["constraint_ledger"]["exists"] is True
    assert metadata["constraint_ledger"]["bytes"] > 0
    assert metadata["constraint_ledger"]["capture_source"] == "file"
    assert metadata["stdout"]["exists"] is True


def test_run_evidence_detects_openclaw_final_visible_text_marker(tmp_path):
    stdout_text = json.dumps(
        {
            "finalAssistantVisibleText": "FINAL REPORT:\nRecovered result.",
            "completion": {"stopReason": "stop"},
        }
    )

    evidence = build_run_evidence(
        worker={"worker_id": "wrk_openclaw_final", "profile": "openclaw-general", "execution_mode": "docker"},
        run_id="run_openclaw_final",
        runtime_name="openclaw",
        model="openai/gpt-5.2",
        command=["openclaw"],
        env={},
        workspace_dir=tmp_path,
        stdout_text=stdout_text,
        stderr_text="",
        output_text="Recovered result.",
        error_text="",
        exit_code=0,
        timeout_seconds=None,
        stop_reason="process_exit",
        constraint_ledger=None,
    )

    assert evidence["final_output"]["has_final_report"] is True
    assert evidence["completion_compliance"]["status"] == "pass"


def test_pdf_render_sample_uses_temp_output_without_orphaning_artifact(tmp_path, monkeypatch):
    output_dir = tmp_path / "output"
    output_dir.mkdir()
    (output_dir / "report.pdf").write_bytes(b"%PDF-1.7\n% synthetic\n")

    def fake_run(command, *, check, capture_output, text, timeout):
        _ = check, capture_output, text, timeout
        Path(command[-1] + ".png").write_bytes(b"png")

        class Completed:
            returncode = 0
            stdout = "ok"
            stderr = ""

        return Completed()

    monkeypatch.setattr("workers_projects_runtime.run_evidence.shutil.which", lambda name: "/usr/bin/pdftoppm" if name == "pdftoppm" else None)
    monkeypatch.setattr("workers_projects_runtime.run_evidence.subprocess.run", fake_run)

    evidence = build_run_evidence(
        worker={"worker_id": "wrk_pdf", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_pdf",
        runtime_name="codex-cli",
        model="gpt-test",
        command=["codex", "exec", "Render PDF."],
        env={},
        workspace_dir=tmp_path,
        stdout_text="FINAL REPORT:\nDone",
        stderr_text="debug image path /Users/example/private-workspace/tmp/preview.png",
        output_text="Done",
        error_text="",
        exit_code=0,
        timeout_seconds=None,
        stop_reason="process_exit",
        constraint_ledger=None,
    )

    pdf = next(item for item in evidence["artifacts"]["items"] if item["path"] == "output/report.pdf")
    assert pdf["document_validation"]["render_sample"]["status"] == "pass"
    assert pdf["document_validation"]["render_sample"]["sample_created"] is True
    assert list(output_dir.glob("*glasshive-sample*.png")) == []


def test_content_hygiene_is_generic_and_advisory():
    clean = {"risk_notes": "VAR analysis, VaR limits, and VaR estimate = 0.05 are normal finance prose."}
    negative_context = {
        "quality_notes": "CSV cells are compact and exclude webpage navigation, cookie notices, scripts, and scraped page text."
    }
    long_report = {"output/report.md": "This is ordinary report prose with no crawl or navigation contamination. " * 25}
    long_structured = {"output/firms.csv": "x" * 901}
    dirty = {
        "one": "Please enable JavaScript to continue reading this article.",
        "two": "Subscribe to continue reading the full report.",
        "three": "window.dataLayer = window.dataLayer || [];",
        "four": "Cookie Settings Privacy Policy Terms of Use",
        "five": "Company&nbsp;overview &amp; leadership",
    }

    clean_result = check_content_hygiene(clean)
    negative_context_result = check_content_hygiene(negative_context)
    long_report_result = check_content_hygiene(long_report)
    long_structured_result = check_content_hygiene(long_structured)
    dirty_result = check_content_hygiene(dirty)

    assert clean_result["status"] == "pass"
    assert clean_result["failure_count"] == 0
    assert negative_context_result["status"] == "pass"
    assert long_report_result["status"] == "pass"
    assert long_structured_result["status"] == "warn"
    assert dirty_result["status"] == "warn"
    assert dirty_result["failure_count"] >= 5


def test_run_evidence_derives_backend_from_profile_not_stale_worker_field(tmp_path):
    worker = {
        "worker_id": "wrk_backend_truth",
        "profile": "codex-cli",
        "execution_mode": "host",
        "backend": "openclaw",
    }
    ledger = build_constraint_ledger(instruction="Do the work.", worker=worker, run_id="run_backend_truth")
    evidence = build_run_evidence(
        worker=worker,
        run_id="run_backend_truth",
        runtime_name="codex-cli",
        model="gpt-test",
        command=["codex", "exec", "Do the work."],
        env={},
        workspace_dir=tmp_path,
        stdout_text="FINAL REPORT:\nDone",
        stderr_text="",
        output_text="Done",
        error_text="",
        exit_code=0,
        timeout_seconds=None,
        stop_reason="process_exit",
        constraint_ledger=ledger,
    )

    assert ledger["worker"]["backend"] == "codex-cli"
    assert evidence["worker"]["backend"] == "codex-cli"
    assert derive_legacy_backend_label(profile="claude-code", backend="openclaw") == "claude-code"
    assert derive_legacy_backend_label(profile="custom-profile", runtime="custom-runtime", backend="openclaw") == "custom-runtime"


def test_run_evidence_records_effective_effort_from_command_and_env(tmp_path):
    codex = build_run_evidence(
        worker={
            "worker_id": "wrk_effort",
            "profile": "codex-cli",
            "execution_mode": "host",
            "_effort_projection": {
                "requested": "xhigh",
                "effective": "medium",
                "allowed": ["high", "low", "medium"],
                "route_proven": False,
                "fallback_reason": "xhigh_route_not_proven",
            },
        },
        run_id="run_effort_codex",
        runtime_name="codex-cli",
        model="gpt-test",
        command=["codex", "exec", "-c", 'model_reasoning_effort="medium"', "Do it."],
        env={},
        workspace_dir=tmp_path,
        stdout_text="FINAL REPORT:\nDone",
        stderr_text="",
        output_text="Done",
        error_text="",
        exit_code=0,
        timeout_seconds=None,
        stop_reason="process_exit",
        constraint_ledger=None,
    )
    claude = build_run_evidence(
        worker={"worker_id": "wrk_effort", "profile": "claude-code", "execution_mode": "host"},
        run_id="run_effort_claude",
        runtime_name="claude-code",
        model="claude-test",
        command=["claude", "-p", "--effort", "max", "Do it."],
        env={"WPR_CLAUDE_CODE_EFFORT": "max"},
        workspace_dir=tmp_path,
        stdout_text="FINAL REPORT:\nDone",
        stderr_text="",
        output_text="Done",
        error_text="",
        exit_code=0,
        timeout_seconds=None,
        stop_reason="process_exit",
        constraint_ledger=None,
    )

    assert codex["effort"] == "medium"
    assert codex["effort_projection"] == {
        "requested": "xhigh",
        "effective": "medium",
        "allowed": ["high", "low", "medium"],
        "route_proven": False,
        "fallback_reason": "xhigh_route_not_proven",
    }
    assert claude["effort"] == "max"
    assert claude["effort_projection"]["effective"] == "max"


def test_run_evidence_redacts_long_command_prompt_and_local_command_paths(tmp_path):
    long_prompt = "Private benchmark prompt\n" + ("sensitive details " * 80)
    evidence = build_run_evidence(
        worker={"worker_id": "wrk_redact", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_redact",
        runtime_name="codex-cli",
        model="gpt-test",
        command=["/Users/example/bin/codex", "exec", "-C", "/Users/example/private-workspace", long_prompt],
        env={},
        workspace_dir=tmp_path,
        stdout_text="FINAL REPORT:\nDone",
        stderr_text="debug image path /Users/example/private-workspace/tmp/preview.png",
        output_text="Done",
        error_text="",
        exit_code=0,
        timeout_seconds=None,
        stop_reason="process_exit",
        constraint_ledger=None,
    )

    argv = evidence["command"]["argv_redacted"]
    display = evidence["command"]["display_redacted"]

    assert argv[0] == "codex"
    assert argv[3] == "private-workspace"
    assert argv[-1].startswith("[REDACTED_LONG_ARG chars=")
    assert "sensitive details" not in display
    assert "/Users/example" not in display
    assert "~/bin" not in display
    assert "[REDACTED_LOCAL_PATH]" in evidence["transcript"]["stderr_tail"]
    assert "/Users/example" not in evidence["transcript"]["stderr_tail"]


def test_run_evidence_final_report_ignores_instruction_only_transcript_marker(tmp_path):
    evidence = build_run_evidence(
        worker={"worker_id": "wrk_final_marker", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_final_marker",
        runtime_name="codex-cli",
        model="gpt-test",
        command=["codex", "exec", "Do it."],
        env={},
        workspace_dir=tmp_path,
        stdout_text="The instructions say to end with FINAL REPORT:, but no final answer was captured.",
        stderr_text="",
        output_text="",
        error_text="Interrupted by operator",
        exit_code=None,
        timeout_seconds=None,
        stop_reason="interrupted",
        constraint_ledger=None,
    )

    assert evidence["final_output"]["has_final_report"] is False
    assert evidence["final_output"]["status"] == "failed"


def test_run_evidence_final_report_ignores_structured_progress_marker(tmp_path):
    stdout_text = json.dumps(
        {
            "type": "result",
            "subtype": "interrupted",
            "result": "Progress note: instructions say to end with FINAL REPORT:, but no final answer was captured.",
        }
    )

    evidence = build_run_evidence(
        worker={"worker_id": "wrk_structured_final_marker", "profile": "claude-code", "execution_mode": "host"},
        run_id="run_structured_final_marker",
        runtime_name="claude-code",
        model="claude-test",
        command=["claude", "-p", "Do it."],
        env={},
        workspace_dir=tmp_path,
        stdout_text=stdout_text,
        stderr_text="",
        output_text="Status: no FINAL REPORT: was captured.",
        error_text="Interrupted by timeout",
        exit_code=None,
        timeout_seconds=300,
        stop_reason="timeout",
        constraint_ledger=None,
    )

    assert evidence["final_output"]["has_final_report"] is False
    assert evidence["evidence_result"]["status"] == "fail"


def test_run_evidence_accepts_plain_stdout_final_report(tmp_path):
    stdout_text = (
        "worker progress before final output\n"
        "FINAL REPORT:\n"
        "Created and visually verified the requested PDF report.\n"
    )

    evidence = build_run_evidence(
        worker={"worker_id": "wrk_plain_stdout_final", "profile": "codex-cli", "execution_mode": "docker"},
        run_id="run_plain_stdout_final",
        runtime_name="codex-cli",
        model="gpt-test",
        command=["codex", "exec", "Do it."],
        env={},
        workspace_dir=tmp_path,
        stdout_text=stdout_text,
        stderr_text="",
        output_text="",
        error_text="",
        exit_code=0,
        timeout_seconds=7200,
        stop_reason="process",
        constraint_ledger=None,
    )

    assert evidence["final_output"]["has_final_report"] is True
    assert evidence["completion_compliance"]["status"] == "pass"
    assert evidence["evidence_result"]["status"] == "pass"


def test_run_evidence_classifies_structured_provider_rate_limit(tmp_path):
    stdout_text = json.dumps(
        {
            "type": "result",
            "subtype": "success",
            "is_error": True,
            "api_error_status": 429,
            "result": "You've hit your session limit; resets later.",
        }
    )

    evidence = build_run_evidence(
        worker={"worker_id": "wrk_provider_429", "profile": "claude-code", "execution_mode": "host"},
        run_id="run_provider_429",
        runtime_name="claude-code",
        model="claude-test",
        command=["claude", "-p", "--output-format", "json"],
        env={},
        workspace_dir=tmp_path,
        stdout_text=stdout_text + "\n",
        stderr_text="",
        output_text="",
        error_text="claude-code exited with code 1",
        exit_code=1,
        timeout_seconds=None,
        stop_reason="process_exit",
        constraint_ledger=None,
    )

    classification = evidence["failure_classification"]
    assert classification["failure_class"] == "provider_rate_limited"
    assert classification["retryable"] is True
    assert "workspace_continue" in classification["recommended_recovery"]
    assert "api_error_status: 429" in classification["diagnostic_summary"]
    assert any(
        reason.get("failure_class") == "provider_rate_limited"
        for reason in evidence["evidence_result"]["failure_reasons"]
    )


def test_run_evidence_classifies_structured_provider_overload(tmp_path):
    stdout_text = json.dumps(
        {
            "type": "result",
            "subtype": "success",
            "is_error": True,
            "api_error_status": 529,
            "result": "API Error: 529 Overloaded. This is usually temporary.",
        }
    )

    evidence = build_run_evidence(
        worker={"worker_id": "wrk_provider_529", "profile": "claude-code", "execution_mode": "host"},
        run_id="run_provider_529",
        runtime_name="claude-code",
        model="claude-test",
        command=["claude", "-p", "--output-format", "json"],
        env={},
        workspace_dir=tmp_path,
        stdout_text=stdout_text + "\n",
        stderr_text="",
        output_text="",
        error_text="claude-code exited with code 1",
        exit_code=1,
        timeout_seconds=None,
        stop_reason="process_exit",
        constraint_ledger=None,
    )

    classification = evidence["failure_classification"]
    assert classification["failure_class"] == "provider_response_failed"
    assert classification["retryable"] is True
    assert "api_error_status: 529" in classification["diagnostic_summary"]


def test_run_evidence_classifies_structured_provider_content_filter(tmp_path):
    stdout_text = json.dumps(
        {
            "type": "turn.failed",
            "error": {
                "code": "content_filter",
                "message": "content_filter",
            },
        }
    )

    evidence = build_run_evidence(
        worker={"worker_id": "wrk_provider_filter", "profile": "claude-code", "execution_mode": "host"},
        run_id="run_provider_filter",
        runtime_name="claude-code",
        model="claude-test",
        command=["claude", "-p", "--output-format", "json"],
        env={},
        workspace_dir=tmp_path,
        stdout_text=stdout_text + "\n",
        stderr_text="",
        output_text="",
        error_text="claude-code exited with code 1",
        exit_code=1,
        timeout_seconds=None,
        stop_reason="process_exit",
        constraint_ledger=None,
    )

    classification = evidence["failure_classification"]
    assert classification["failure_class"] == "provider_content_filter"
    assert classification["retryable"] is False
    assert any(
        reason.get("failure_class") == "provider_content_filter"
        for reason in evidence["evidence_result"]["failure_reasons"]
    )


def test_run_evidence_classifies_structured_provider_auth_missing(tmp_path):
    stdout_text = json.dumps(
        {
            "type": "result",
            "subtype": "success",
            "is_error": True,
            "api_error_status": 401,
            "result": "Not logged in. Please run /login.",
        }
    )

    evidence = build_run_evidence(
        worker={"worker_id": "wrk_provider_auth", "profile": "claude-code", "execution_mode": "host"},
        run_id="run_provider_auth",
        runtime_name="claude-code",
        model="claude-test",
        command=["claude", "-p", "--output-format", "json"],
        env={},
        workspace_dir=tmp_path,
        stdout_text=stdout_text + "\n",
        stderr_text="",
        output_text="",
        error_text="claude-code exited with code 1",
        exit_code=1,
        timeout_seconds=None,
        stop_reason="process_exit",
        constraint_ledger=None,
    )

    classification = evidence["failure_classification"]
    assert classification["failure_class"] == "provider_auth_missing"
    assert classification["retryable"] is False
    assert "api_error_status: 401" in classification["diagnostic_summary"]


def test_run_evidence_still_classifies_structured_provider_403_auth_missing(tmp_path):
    stdout_text = json.dumps(
        {
            "type": "result",
            "subtype": "error",
            "is_error": True,
            "api_error_status": 403,
            "result": "403 Forbidden: provider credentials are not authorized.",
        }
    )

    evidence = build_run_evidence(
        worker={"worker_id": "wrk_provider_403", "profile": "claude-code", "execution_mode": "host"},
        run_id="run_provider_403",
        runtime_name="claude-code",
        model="claude-test",
        command=["claude", "-p", "--output-format", "json"],
        env={},
        workspace_dir=tmp_path,
        stdout_text=stdout_text + "\n",
        stderr_text="",
        output_text="",
        error_text="claude-code exited with code 1",
        exit_code=1,
        timeout_seconds=None,
        stop_reason="process_exit",
        constraint_ledger=None,
    )

    classification = evidence["failure_classification"]
    assert classification["failure_class"] == "provider_auth_missing"
    assert classification["retryable"] is False
    assert "api_error_status: 403" in classification["diagnostic_summary"]


def test_run_evidence_classifies_structured_provider_request_rejected(tmp_path):
    stdout_text = json.dumps(
        {
            "type": "result",
            "subtype": "error",
            "is_error": True,
            "api_error_status": 400,
            "error_code": "invalid_request_error",
            "result": "400 Bad Request: request rejected because the route does not support this option.",
        }
    )

    evidence = build_run_evidence(
        worker={"worker_id": "wrk_provider_400", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_provider_400",
        runtime_name="codex-cli",
        model="gpt-test",
        command=["codex", "exec"],
        env={},
        workspace_dir=tmp_path,
        stdout_text=stdout_text + "\n",
        stderr_text="",
        output_text="",
        error_text="codex-cli exited with code 1",
        exit_code=1,
        timeout_seconds=None,
        stop_reason="process_exit",
        constraint_ledger=None,
    )

    classification = evidence["failure_classification"]
    assert classification["failure_class"] == "provider_request_rejected"
    assert classification["retryable"] is False
    assert "unsupported option" in classification["recommended_recovery"]


def test_run_evidence_classifies_zero_exit_structured_provider_error(tmp_path):
    stdout_text = json.dumps(
        {
            "type": "result",
            "subtype": "success",
            "is_error": True,
            "api_error_status": 429,
            "result": "Too Many Requests",
        }
    )

    evidence = build_run_evidence(
        worker={"worker_id": "wrk_provider_zero_exit", "profile": "claude-code", "execution_mode": "host"},
        run_id="run_provider_zero_exit",
        runtime_name="claude-code",
        model="claude-test",
        command=["claude", "-p", "--output-format", "json"],
        env={},
        workspace_dir=tmp_path,
        stdout_text=stdout_text + "\n",
        stderr_text="",
        output_text="",
        error_text="",
        exit_code=0,
        timeout_seconds=None,
        stop_reason="process_exit",
        constraint_ledger=None,
    )

    classification = evidence["failure_classification"]
    assert classification["failure_class"] == "provider_rate_limited"
    assert classification["retryable"] is True
    assert any(
        reason.get("failure_class") == "provider_rate_limited"
        for reason in evidence["evidence_result"]["failure_reasons"]
    )


def test_run_evidence_does_not_classify_benign_provider_terms_in_success_output(tmp_path):
    stdout_text = json.dumps(
        {
            "type": "item.completed",
            "item": {
                "type": "agent_message",
                "text": (
                    "FINAL REPORT:\n"
                    "The research mentions rate limit, HTTP 429, and content filter topics as domain facts, "
                    "not as provider failures."
                ),
            },
        }
    )

    evidence = build_run_evidence(
        worker={"worker_id": "wrk_provider_terms", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_provider_terms",
        runtime_name="codex-cli",
        model="gpt-test",
        command=["codex", "exec"],
        env={},
        workspace_dir=tmp_path,
        stdout_text=stdout_text + "\n",
        stderr_text="",
        output_text=(
            "FINAL REPORT:\n"
            "The research mentions rate limit, HTTP 429, and content filter topics as domain facts."
        ),
        error_text="",
        exit_code=0,
        timeout_seconds=None,
        stop_reason="process_exit",
        constraint_ledger=None,
    )

    assert evidence["failure_classification"]["status"] == "not_applicable"
    assert evidence["evidence_result"]["status"] == "pass"


def test_run_evidence_does_not_classify_false_structured_error_fields(tmp_path):
    stdout_text = json.dumps(
        {
            "type": "result",
            "subtype": "success",
            "is_error": "false",
            "error": "none",
            "failure": "false",
            "result": "FINAL REPORT:\nCompleted successfully.",
        }
    )

    evidence = build_run_evidence(
        worker={"worker_id": "wrk_false_error_fields", "profile": "claude-code", "execution_mode": "host"},
        run_id="run_false_error_fields",
        runtime_name="claude-code",
        model="claude-test",
        command=["claude", "-p", "--output-format", "json"],
        env={},
        workspace_dir=tmp_path,
        stdout_text=stdout_text + "\n",
        stderr_text="",
        output_text="FINAL REPORT:\nCompleted successfully.",
        error_text="",
        exit_code=0,
        timeout_seconds=None,
        stop_reason="process_exit",
        constraint_ledger=None,
    )

    assert evidence["failure_classification"]["status"] == "not_applicable"
    assert evidence["evidence_result"]["status"] == "pass"


def test_run_evidence_does_not_treat_browser_snapshot_node_ids_as_provider_status(tmp_path):
    stdout_text = json.dumps(
        {
            "type": "result",
            "subtype": "success",
            "is_error": True,
            "result": (
                "browser accessibility snapshot\n"
                "401 menu item\n"
                "403 close button\n"
                "FINAL REPORT:\nChrome is open on the requested page."
            ),
        }
    )

    evidence = build_run_evidence(
        worker={"worker_id": "wrk_browser_snapshot_ids", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_browser_snapshot_ids",
        runtime_name="codex-cli",
        model="gpt-test",
        command=["codex", "exec"],
        env={},
        workspace_dir=tmp_path,
        stdout_text=stdout_text + "\n",
        stderr_text="",
        output_text="FINAL REPORT:\nChrome is open on the requested page.",
        error_text="",
        exit_code=0,
        timeout_seconds=None,
        stop_reason="process_exit",
        constraint_ledger=None,
    )

    classification = evidence["failure_classification"]
    assert classification["status"] == "not_applicable"
    assert "provider_auth_missing" not in json.dumps(evidence)
    assert evidence["evidence_result"]["status"] == "pass"


@pytest.mark.parametrize(
    "status_like_text, forbidden_failure_class",
    [
        ("400 Bad Request banner in the page body", "provider_request_rejected"),
        ("429 Too Many Requests heading on the site", "provider_rate_limited"),
        ("503 Service Unavailable page from the target site", "provider_response_failed"),
        ("529 overloaded label in a browser snapshot", "provider_response_failed"),
    ],
)
def test_run_evidence_success_final_report_ignores_browser_page_status_like_text(
    tmp_path,
    status_like_text,
    forbidden_failure_class,
):
    stdout_text = json.dumps(
        {
            "type": "result",
            "subtype": "success",
            "is_error": True,
            "result": (
                "browser accessibility snapshot\n"
                f"{status_like_text}\n"
                "FINAL REPORT:\nThe browser state was inspected and the task is complete."
            ),
        }
    )

    evidence = build_run_evidence(
        worker={"worker_id": "wrk_browser_status_text", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_browser_status_text",
        runtime_name="codex-cli",
        model="gpt-test",
        command=["codex", "exec"],
        env={},
        workspace_dir=tmp_path,
        stdout_text=stdout_text + "\n",
        stderr_text="",
        output_text="FINAL REPORT:\nThe browser state was inspected and the task is complete.",
        error_text="",
        exit_code=0,
        timeout_seconds=None,
        stop_reason="process_exit",
        constraint_ledger=None,
    )

    assert evidence["failure_classification"]["status"] == "not_applicable"
    assert forbidden_failure_class not in json.dumps(evidence)
    assert evidence["evidence_result"]["status"] == "pass"


def test_run_evidence_does_not_classify_failed_agent_substep_after_final_report(tmp_path):
    stdout_text = "\n".join(
        [
            json.dumps(
                {
                    "type": "item.completed",
                    "item": {
                        "type": "command_execution",
                        "command": "ls missing-path",
                        "aggregated_output": "ls: missing-path: No such file or directory",
                        "exit_code": 2,
                        "status": "failed",
                    },
                }
            ),
            json.dumps(
                {
                    "type": "item.completed",
                    "item": {
                        "type": "agent_message",
                        "text": "FINAL REPORT:\nRecovered from the probe and completed successfully.",
                    },
                }
            ),
        ]
    )

    evidence = build_run_evidence(
        worker={"worker_id": "wrk_failed_substep", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_failed_substep",
        runtime_name="codex-cli",
        model="gpt-test",
        command=["codex", "exec"],
        env={},
        workspace_dir=tmp_path,
        stdout_text=stdout_text + "\n",
        stderr_text="",
        output_text="FINAL REPORT:\nRecovered from the probe and completed successfully.",
        error_text="",
        exit_code=0,
        timeout_seconds=None,
        stop_reason="process_exit",
        constraint_ledger=None,
    )

    assert evidence["failure_classification"]["status"] == "not_applicable"
    assert evidence["evidence_result"]["status"] == "pass"


def test_run_evidence_accepts_markdown_decorated_final_report_marker(tmp_path):
    evidence = build_run_evidence(
        worker={"worker_id": "wrk_decorated_final", "profile": "claude-code", "execution_mode": "host"},
        run_id="run_decorated_final",
        runtime_name="claude-code",
        model="claude-test",
        command=["claude", "-p"],
        env={},
        workspace_dir=tmp_path,
        stdout_text=json.dumps(
            {
                "type": "result",
                "subtype": "success",
                "result": "### FINAL REPORT:\nDone.",
            }
        ),
        stderr_text="",
        output_text="**FINAL REPORT:**\nDone.",
        error_text="",
        exit_code=0,
        timeout_seconds=None,
        stop_reason="process_exit",
        constraint_ledger=None,
    )

    assert evidence["final_output"]["has_final_report"] is True
    assert evidence["evidence_result"]["status"] == "pass"


def test_run_evidence_detects_claude_result_final_report_marker(tmp_path):
    stdout_text = json.dumps(
        {
            "type": "result",
            "subtype": "success",
            "result": "All files verified.\n\nFINAL REPORT:\nCreated output/summary.csv.",
        }
    )

    evidence = build_run_evidence(
        worker={"worker_id": "wrk_claude_final", "profile": "claude-code", "execution_mode": "host"},
        run_id="run_claude_final",
        runtime_name="claude-code",
        model="claude-test",
        command=["claude", "-p", "--effort", "max", "Do it."],
        env={"WPR_CLAUDE_CODE_EFFORT": "max"},
        workspace_dir=tmp_path,
        stdout_text=stdout_text,
        stderr_text="",
        output_text="Created output/summary.csv.",
        error_text="",
        exit_code=0,
        timeout_seconds=None,
        stop_reason="process_exit",
        constraint_ledger=None,
    )

    assert evidence["final_output"]["has_final_report"] is True
    assert evidence["completion_compliance"]["status"] == "pass"


def test_run_evidence_detects_stdout_final_report_when_output_text_missing(tmp_path):
    stdout_text = json.dumps(
        {
            "type": "item.completed",
            "item": {
                "type": "agent_message",
                "text": "FINAL REPORT:\nCreated output/report.md.",
            },
        }
    )

    evidence = build_run_evidence(
        worker={"worker_id": "wrk_stdout_final", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_stdout_final",
        runtime_name="codex-cli",
        model="gpt-test",
        command=["codex", "exec", "Do it."],
        env={},
        workspace_dir=tmp_path,
        stdout_text=stdout_text,
        stderr_text="",
        output_text="",
        error_text="",
        exit_code=0,
        timeout_seconds=None,
        stop_reason="process_exit",
        constraint_ledger=None,
    )

    assert evidence["final_output"]["has_final_report"] is True


def test_constraint_compliance_avoids_generic_future_and_approximate_false_positives(tmp_path):
    output_dir = tmp_path / "output"
    output_dir.mkdir()
    (output_dir / "report.md").write_text(
        "The project plans a March 2027 launch and approximately 5% growth.\n"
        "The analysis cites official sources through May 2026.\n"
    )
    research_dir = tmp_path / "research"
    research_dir.mkdir()
    (research_dir / "notes.md").write_text("Best effort given available public sources.\n")
    ledger = build_constraint_ledger(
        instruction="Use sources from January 2024 through May 2026 only.",
        worker={"worker_id": "wrk_constraints", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_constraints",
    )

    evidence = build_run_evidence(
        worker={"worker_id": "wrk_constraints", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_constraints",
        runtime_name="codex-cli",
        model="gpt-test",
        command=["codex", "exec", "Analyze."],
        env={},
        workspace_dir=tmp_path,
        stdout_text="FINAL REPORT:\nDone",
        stderr_text="",
        output_text="Done",
        error_text="",
        exit_code=0,
        timeout_seconds=300,
        stop_reason="process_exit",
        constraint_ledger=ledger,
    )

    assert evidence["timeout"]["exit_source"] == "process"
    assert evidence["constraint_compliance"]["status"] == "pass"


def test_constraint_compliance_fails_when_planning_file_omits_strict_source_window(tmp_path):
    output_dir = tmp_path / "output"
    output_dir.mkdir()
    (output_dir / "report.md").write_text("FINAL REPORT:\nThe final answer used in-window sources.\n")
    specs_dir = tmp_path / "specs"
    specs_dir.mkdir()
    (specs_dir / "SPEC.md").write_text("Plan: gather public evidence and write the report.\n")
    ledger = build_constraint_ledger(
        instruction="Use sources from January 2024 through May 2026 only.",
        worker={"worker_id": "wrk_spec_warning", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_spec_warning",
    )

    evidence = build_run_evidence(
        worker={"worker_id": "wrk_spec_warning", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_spec_warning",
        runtime_name="codex-cli",
        model="gpt-test",
        command=["codex", "exec", "Analyze."],
        env={},
        workspace_dir=tmp_path,
        stdout_text="FINAL REPORT:\nDone",
        stderr_text="",
        output_text="FINAL REPORT:\nDone",
        error_text="",
        exit_code=0,
        timeout_seconds=300,
        stop_reason="process_exit",
        constraint_ledger=ledger,
    )

    assert evidence["constraint_compliance"]["status"] == "fail"
    assert any(
        issue["reason"] == "strict source/date constraints not referenced in planning file"
        for issue in evidence["constraint_compliance"]["issues"]
    )
    assert evidence["evidence_result"]["status"] == "fail"


def test_constraint_compliance_allows_final_answer_only_scheduled_prompt_operational_evidence(tmp_path):
    instruction = (
        "You are executing a Prompt Workbench scheduled prompt.\n"
        "Use the rendered prompt and provided snapshot files as the source of truth.\n"
        "Memory write mode: off.\n"
        "If memory write mode is off, do not modify account memory.\n"
        "If memory write mode is propose, write governed memory proposals only.\n"
        "For memory proposals, write UTF-8 JSON under the scratchpad folder named memory-proposals-yyyymmddHHmm.json with an actions array of set/delete objects.\n"
        "Do not inspect or modify private memories, conversations, files, email, or accounts.\n"
        "Capture evidence of the completed scheduled prompt only in the final report.\n"
        "Report back in the final answer only.\n"
        "End with exactly: FINAL REPORT: QA_OK QA_WORKBENCH_EXECUTOR_SYNTHETIC"
    )
    (tmp_path / "prompt.md").write_text(instruction)
    scheduled_prompt_dir = tmp_path / "scheduled-prompt"
    scheduled_prompt_dir.mkdir()
    (scheduled_prompt_dir / "utf8_static_server.py").write_text(
        "# Infrastructure preview helper, not a user deliverable.\n"
        "COMMENT = 'strict constraints should be followed wherever possible'\n"
    )
    assert not is_user_deliverable_relative_path("scheduled-prompt/utf8_static_server.py")
    worker = {
        "worker_id": "wrk_scheduled_prompt",
        "profile": "codex-cli",
        "execution_mode": "host",
        "bootstrap_bundle_json": json.dumps(
            {
                "files": [
                    {"path": "scheduled-prompt/rendered-prompt.md"},
                    {"path": "scheduled-prompt/variable-snapshot.json"},
                    {"path": "scheduled-prompt/run-contract.md"},
                ]
            }
        ),
    }
    ledger = build_constraint_ledger(
        instruction=instruction,
        worker=worker,
        run_id="run_scheduled_prompt",
    )

    evidence = build_run_evidence(
        worker=worker,
        run_id="run_scheduled_prompt",
        runtime_name="codex-cli",
        model="gpt-test",
        command=["codex", "exec", instruction],
        env={},
        workspace_dir=tmp_path,
        stdout_text="FINAL REPORT: QA_OK QA_WORKBENCH_EXECUTOR_SYNTHETIC",
        stderr_text="",
        output_text="FINAL REPORT: QA_OK QA_WORKBENCH_EXECUTOR_SYNTHETIC",
        error_text="",
        exit_code=0,
        timeout_seconds=300,
        stop_reason="process_exit",
        constraint_ledger=ledger,
    )

    assert evidence["constraint_compliance"]["status"] == "pass"
    assert evidence["evidence_result"]["status"] == "pass"


def test_completion_compliance_counts_scheduled_prompt_private_scratchpad_artifacts(tmp_path):
    my_folder = tmp_path / "private-my-folder"
    my_folder.mkdir()
    scratchpad = my_folder / "202606250300.md"
    proposals = my_folder / "memory-proposals-202606250300.json"
    scratchpad.write_text("FINAL REPORT:\nSynthetic scheduled prompt notes.\n", encoding="utf-8")
    proposals.write_text('{"actions":[{"action":"set","key":"working","value":"synthetic"}]}', encoding="utf-8")
    scheduled_prompt_dir = tmp_path / "scheduled-prompt"
    scheduled_prompt_dir.mkdir()
    (scheduled_prompt_dir / "variable-snapshot.json").write_text(
        json.dumps(
            {
                "items": [
                    {
                        "placeholder": "local.viventium.my_folder",
                        "wrapper": "local.viventium.my_folder",
                        "kind": "path",
                        "value": str(my_folder),
                    }
                ]
            }
        ),
        encoding="utf-8",
    )
    instruction = (
        "You are executing a Prompt Workbench scheduled prompt.\n"
        "Memory write mode: propose.\n"
        "Write a markdown private scratchpad file named yyyymmddHHmm.md.\n"
        "For memory proposals, write UTF-8 JSON under that folder named "
        "memory-proposals-yyyymmddHHmm.json with an actions array of set/delete objects.\n"
        "End every run with a concise `FINAL REPORT:` section."
    )
    ledger = build_constraint_ledger(
        instruction=instruction,
        worker={"worker_id": "wrk_scheduled_prompt_private_artifacts", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_scheduled_prompt_private_artifacts",
    )

    evidence = build_run_evidence(
        worker={"worker_id": "wrk_scheduled_prompt_private_artifacts", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_scheduled_prompt_private_artifacts",
        runtime_name="codex-cli",
        model="gpt-test",
        command=["codex", "exec", instruction],
        env={},
        workspace_dir=tmp_path,
        stdout_text="FINAL REPORT:\nCompleted synthetic scheduled prompt.",
        stderr_text="",
        output_text="FINAL REPORT:\nCompleted synthetic scheduled prompt.",
        error_text="",
        exit_code=0,
        timeout_seconds=300,
        stop_reason="process_exit",
        constraint_ledger=ledger,
    )

    completion = evidence["completion_compliance"]
    assert set(completion["required_artifact_types"]) == {"json", "md"}
    assert completion["missing_required_artifact_types"] == []
    assert completion["status"] == "pass"
    assert "private-scratchpad/202606250300.md" in completion["deliverable_artifact_paths"]
    assert "private-scratchpad/memory-proposals-202606250300.json" in completion["deliverable_artifact_paths"]
    assert str(my_folder) not in json.dumps(evidence["artifacts"])
    assert evidence["evidence_result"]["status"] == "pass"


def test_constraint_compliance_allows_user_facing_root_plan_deliverable_without_ledger_link(tmp_path):
    (tmp_path / "research-implementation-plan.md").write_text(
        "# Research Implementation Plan\n\n"
        "FINAL REPORT:\n"
        "This plan uses cited in-window sources and is the requested user-facing markdown deliverable.\n"
    )
    ledger = build_constraint_ledger(
        instruction=(
            "Use sources from January 2024 through May 2026 only. "
            "Produce a markdown implementation plan in the project folder."
        ),
        worker={"worker_id": "wrk_root_plan", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_root_plan",
    )

    evidence = build_run_evidence(
        worker={"worker_id": "wrk_root_plan", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_root_plan",
        runtime_name="codex-cli",
        model="gpt-test",
        command=["codex", "exec", "Analyze."],
        env={},
        workspace_dir=tmp_path,
        stdout_text="FINAL REPORT:\nCreated the plan.",
        stderr_text="",
        output_text="FINAL REPORT:\nCreated the plan.",
        error_text="",
        exit_code=0,
        timeout_seconds=300,
        stop_reason="process_exit",
        constraint_ledger=ledger,
    )

    assert evidence["constraint_compliance"]["status"] == "pass"
    assert evidence["evidence_result"]["status"] == "pass"


@pytest.mark.parametrize(
    "relative_path",
    [
        "output/research-implementation-plan.md",
        "artifacts/research-plan.md",
        "reports/research-spec.md",
    ],
)
def test_constraint_compliance_allows_user_facing_deliverable_roots_named_plan_or_spec(tmp_path, relative_path):
    artifact_path = tmp_path / relative_path
    artifact_path.parent.mkdir(parents=True, exist_ok=True)
    artifact_path.write_text(
        "# Research Deliverable\n\n"
        "FINAL REPORT:\n"
        "This requested deliverable uses cited in-window sources.\n"
    )
    ledger = build_constraint_ledger(
        instruction=(
            "Use sources from January 2024 through May 2026 only. "
            "Produce a markdown implementation plan in the project folder."
        ),
        worker={"worker_id": "wrk_deliverable_roots", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_deliverable_roots",
    )

    evidence = build_run_evidence(
        worker={"worker_id": "wrk_deliverable_roots", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_deliverable_roots",
        runtime_name="codex-cli",
        model="gpt-test",
        command=["codex", "exec", "Analyze."],
        env={},
        workspace_dir=tmp_path,
        stdout_text="FINAL REPORT:\nCreated the requested deliverable.",
        stderr_text="",
        output_text="FINAL REPORT:\nCreated the requested deliverable.",
        error_text="",
        exit_code=0,
        timeout_seconds=300,
        stop_reason="process_exit",
        constraint_ledger=ledger,
    )

    assert evidence["constraint_compliance"]["status"] == "pass"
    assert evidence["evidence_result"]["status"] == "pass"


def test_constraint_compliance_still_requires_internal_named_root_files_to_preserve_constraints(tmp_path):
    (tmp_path / "delegation-notes.md").write_text(
        "# Delegation Notes\n\n"
        "Ask a helper to research the topic and return findings.\n"
    )
    ledger = build_constraint_ledger(
        instruction="Use sources from January 2024 through May 2026 only.",
        worker={"worker_id": "wrk_internal_named_root", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_internal_named_root",
    )

    evidence = build_run_evidence(
        worker={"worker_id": "wrk_internal_named_root", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_internal_named_root",
        runtime_name="codex-cli",
        model="gpt-test",
        command=["codex", "exec", "Analyze."],
        env={},
        workspace_dir=tmp_path,
        stdout_text="FINAL REPORT:\nDone.",
        stderr_text="",
        output_text="FINAL REPORT:\nDone.",
        error_text="",
        exit_code=0,
        timeout_seconds=300,
        stop_reason="process_exit",
        constraint_ledger=ledger,
    )

    assert evidence["constraint_compliance"]["status"] == "fail"
    assert any(
        issue["reason"] == "strict source/date constraints not referenced in planning file"
        for issue in evidence["constraint_compliance"]["issues"]
    )


def test_constraint_compliance_still_flags_softened_constraints_in_root_plan_deliverable(tmp_path):
    (tmp_path / "research-implementation-plan.md").write_text(
        "# Research Implementation Plan\n\n"
        "Must use sources from January 2024 through May 2026 only wherever possible.\n"
    )
    ledger = build_constraint_ledger(
        instruction="Use sources from January 2024 through May 2026 only.",
        worker={"worker_id": "wrk_root_plan_softened", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_root_plan_softened",
    )

    evidence = build_run_evidence(
        worker={"worker_id": "wrk_root_plan_softened", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_root_plan_softened",
        runtime_name="codex-cli",
        model="gpt-test",
        command=["codex", "exec", "Analyze."],
        env={},
        workspace_dir=tmp_path,
        stdout_text="FINAL REPORT:\nCreated the plan.",
        stderr_text="",
        output_text="FINAL REPORT:\nCreated the plan.",
        error_text="",
        exit_code=0,
        timeout_seconds=300,
        stop_reason="process_exit",
        constraint_ledger=ledger,
    )

    assert evidence["constraint_compliance"]["status"] == "fail"
    assert any(
        issue["reason"] == "strict constraint softened in workspace file"
        for issue in evidence["constraint_compliance"]["issues"]
    )


def test_constraint_compliance_still_scans_root_plan_deliverable_for_source_date_drift(tmp_path):
    (tmp_path / "research-implementation-plan.md").write_text(
        "# Research Implementation Plan\n\n"
        "A cited source published June 2026 supports the plan.\n"
    )
    ledger = build_constraint_ledger(
        instruction=(
            "Use sources from January 2024 through May 2026 only. "
            "Produce a markdown implementation plan in the project folder."
        ),
        worker={"worker_id": "wrk_root_plan_drift", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_root_plan_drift",
    )

    evidence = build_run_evidence(
        worker={"worker_id": "wrk_root_plan_drift", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_root_plan_drift",
        runtime_name="codex-cli",
        model="gpt-test",
        command=["codex", "exec", "Analyze."],
        env={},
        workspace_dir=tmp_path,
        stdout_text="FINAL REPORT:\nCreated the plan.",
        stderr_text="",
        output_text="FINAL REPORT:\nCreated the plan.",
        error_text="",
        exit_code=0,
        timeout_seconds=300,
        stop_reason="process_exit",
        constraint_ledger=ledger,
    )

    assert evidence["constraint_compliance"]["status"] == "fail"
    assert any(
        issue["reason"] == "date/source window widened past ledger limit"
        for issue in evidence["constraint_compliance"]["issues"]
    )


def test_constraint_compliance_scans_final_output_for_source_window_drift(tmp_path):
    ledger = build_constraint_ledger(
        instruction="Use sources from January 2024 through May 2026 only.",
        worker={"worker_id": "wrk_final_window", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_final_window",
    )

    evidence = build_run_evidence(
        worker={"worker_id": "wrk_final_window", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_final_window",
        runtime_name="codex-cli",
        model="gpt-test",
        command=["codex", "exec", "Analyze."],
        env={},
        workspace_dir=tmp_path,
        stdout_text="FINAL REPORT:\nDone",
        stderr_text="",
        output_text="FINAL REPORT:\nUsed a primary source published June 2026 for the score.",
        error_text="",
        exit_code=0,
        timeout_seconds=300,
        stop_reason="process_exit",
        constraint_ledger=ledger,
    )

    assert evidence["constraint_compliance"]["status"] == "fail"
    assert any(issue["path"] == "<final-output>" for issue in evidence["constraint_compliance"]["issues"])


def test_constraint_compliance_allows_access_timestamp_without_widening_source_window(tmp_path):
    ledger = build_constraint_ledger(
        instruction="Use sources from January 2024 through May 2026 only.",
        worker={"worker_id": "wrk_access_date", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_access_date",
    )

    evidence = build_run_evidence(
        worker={"worker_id": "wrk_access_date", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_access_date",
        runtime_name="codex-cli",
        model="gpt-test",
        command=["codex", "exec", "Analyze."],
        env={},
        workspace_dir=tmp_path,
        stdout_text="FINAL REPORT:\nDone",
        stderr_text="",
        output_text="FINAL REPORT:\nSource publication date is March 2026. Access date: 2026-06-22.",
        error_text="",
        exit_code=0,
        timeout_seconds=300,
        stop_reason="process_exit",
        constraint_ledger=ledger,
    )

    assert evidence["constraint_compliance"]["status"] == "pass"


def test_constraint_compliance_allows_explicitly_excluded_out_of_window_note(tmp_path):
    ledger = build_constraint_ledger(
        instruction="Use sources from January 2024 through May 2026 only.",
        worker={"worker_id": "wrk_excluded_date", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_excluded_date",
    )

    evidence = build_run_evidence(
        worker={"worker_id": "wrk_excluded_date", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_excluded_date",
        runtime_name="codex-cli",
        model="gpt-test",
        command=["codex", "exec", "Analyze."],
        env={},
        workspace_dir=tmp_path,
        stdout_text="FINAL REPORT:\nDone",
        stderr_text="",
        output_text=(
            "FINAL REPORT:\n"
            "June 2026 article is OUT-OF-WINDOW and excluded from scoring; "
            "a July 2026 filing is outside January 2024-May 2026 window and not used; "
            "in-window facts use March 2026 sources."
        ),
        error_text="",
        exit_code=0,
        timeout_seconds=300,
        stop_reason="process_exit",
        constraint_ledger=ledger,
    )

    assert evidence["constraint_compliance"]["status"] == "pass"


def test_constraint_compliance_does_not_treat_research_notes_as_planning_softening(tmp_path):
    research_dir = tmp_path / "research"
    research_dir.mkdir()
    (research_dir / "firm-notes.md").write_text(
        "Evidence note: figure is likely blended and should be flagged for human review, "
        "but the final score uses in-window primary sources only.\n"
    )
    ledger = build_constraint_ledger(
        instruction="Use sources from January 2024 through May 2026 only.",
        worker={"worker_id": "wrk_research_note", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_research_note",
    )

    evidence = build_run_evidence(
        worker={"worker_id": "wrk_research_note", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_research_note",
        runtime_name="codex-cli",
        model="gpt-test",
        command=["codex", "exec", "Analyze."],
        env={},
        workspace_dir=tmp_path,
        stdout_text="FINAL REPORT:\nDone",
        stderr_text="",
        output_text="FINAL REPORT:\nDone",
        error_text="",
        exit_code=0,
        timeout_seconds=300,
        stop_reason="process_exit",
        constraint_ledger=ledger,
    )

    assert evidence["constraint_compliance"]["status"] == "pass"


def test_constraint_compliance_scans_xlsx_artifact_text_for_source_window_drift(tmp_path):
    openpyxl = pytest.importorskip("openpyxl")
    output_dir = tmp_path / "output"
    output_dir.mkdir()
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Sources"
    worksheet.append(["Entity", "Source note"])
    worksheet.append(["ExampleCo", "Primary source published June 2026; used for scoring."])
    workbook.save(output_dir / "source-ledger.xlsx")
    (output_dir / "report.md").write_text("FINAL REPORT:\nSee the workbook.\n")
    ledger = build_constraint_ledger(
        instruction="Use sources from January 2024 through May 2026 only.",
        worker={"worker_id": "wrk_xlsx_window", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_xlsx_window",
    )

    evidence = build_run_evidence(
        worker={"worker_id": "wrk_xlsx_window", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_xlsx_window",
        runtime_name="codex-cli",
        model="gpt-test",
        command=["codex", "exec", "Analyze."],
        env={},
        workspace_dir=tmp_path,
        stdout_text="FINAL REPORT:\nDone",
        stderr_text="",
        output_text="FINAL REPORT:\nDone",
        error_text="",
        exit_code=0,
        timeout_seconds=300,
        stop_reason="process_exit",
        constraint_ledger=ledger,
    )

    assert evidence["constraint_compliance"]["status"] == "fail"
    assert any(issue["path"] == "output/source-ledger.xlsx" for issue in evidence["constraint_compliance"]["issues"])
    assert "output/source-ledger.xlsx" in evidence["constraint_compliance"]["scanned_sources"]


def test_constraint_compliance_warns_when_binary_text_extraction_is_unavailable(tmp_path, monkeypatch):
    output_dir = tmp_path / "output"
    output_dir.mkdir()
    (output_dir / "source-ledger.xlsx").write_bytes(b"synthetic workbook bytes")
    ledger = build_constraint_ledger(
        instruction="Use sources from January 2024 through May 2026 only.",
        worker={"worker_id": "wrk_xlsx_unavailable", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_xlsx_unavailable",
    )
    monkeypatch.setattr(
        run_evidence,
        "_extract_xlsx_text",
        lambda _path: ("", "openpyxl unavailable for xlsx constraint scan"),
    )

    evidence = build_run_evidence(
        worker={"worker_id": "wrk_xlsx_unavailable", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_xlsx_unavailable",
        runtime_name="codex-cli",
        model="gpt-5.4",
        command=["codex", "exec", "-"],
        env={},
        workspace_dir=tmp_path,
        stdout_text="FINAL REPORT:\ncreated files",
        stderr_text="",
        output_text="FINAL REPORT:\ncreated files",
        error_text="",
        exit_code=0,
        timeout_seconds=None,
        stop_reason="",
        constraint_ledger=ledger,
    )

    assert evidence["constraint_compliance"]["status"] == "warn"
    assert any(
        issue["path"] == "output/source-ledger.xlsx"
        and issue["reason"] == "constraint text extraction unavailable"
        for issue in evidence["constraint_compliance"]["issues"]
    )


def test_constraint_compliance_fails_when_planning_file_widens_source_window(tmp_path):
    output_dir = tmp_path / "output"
    output_dir.mkdir()
    (output_dir / "report.md").write_text("FINAL REPORT:\nThe final answer used in-window sources.\n")
    research_dir = tmp_path / "research"
    research_dir.mkdir()
    (research_dir / "SPEC.md").write_text(
        "Research plan: use primary sources from January 2024 through June 2026.\n"
        "Deliver the final table after scoring.\n"
    )
    ledger = build_constraint_ledger(
        instruction="Use primary sources from January 2024 through May 2026 only. Deliver a final table.",
        worker={"worker_id": "wrk_plan_widen", "profile": "claude-code", "execution_mode": "host"},
        run_id="run_plan_widen",
    )

    evidence = build_run_evidence(
        worker={"worker_id": "wrk_plan_widen", "profile": "claude-code", "execution_mode": "host"},
        run_id="run_plan_widen",
        runtime_name="claude-code",
        model="claude-test",
        command=["claude", "-p"],
        env={},
        workspace_dir=tmp_path,
        stdout_text='{"type":"result","result":"FINAL REPORT:\\nDone"}',
        stderr_text="",
        output_text="FINAL REPORT:\nDone",
        error_text="",
        exit_code=0,
        timeout_seconds=None,
        stop_reason="process_exit",
        constraint_ledger=ledger,
    )

    assert evidence["constraint_compliance"]["status"] == "fail"
    assert any(
        issue["reason"] == "planning source/date window widened past ledger limit"
        for issue in evidence["constraint_compliance"]["issues"]
    )
    assert evidence["evidence_result"]["status"] == "fail"


def test_constraint_compliance_accepts_planning_file_that_references_ledger(tmp_path):
    output_dir = tmp_path / "output"
    output_dir.mkdir()
    (output_dir / "report.md").write_text("FINAL REPORT:\nThe final answer used in-window sources.\n")
    planning_dir = tmp_path / "planning"
    planning_dir.mkdir()
    (planning_dir / "delegate-plan.md").write_text(
        "Read glasshive-run/constraint-ledger.json before delegating source work.\n"
    )
    ledger = build_constraint_ledger(
        instruction="Use sources from January 2024 through May 2026 only.",
        worker={"worker_id": "wrk_spec_pass", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_spec_pass",
    )

    evidence = build_run_evidence(
        worker={"worker_id": "wrk_spec_pass", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_spec_pass",
        runtime_name="codex-cli",
        model="gpt-test",
        command=["codex", "exec", "Analyze."],
        env={},
        workspace_dir=tmp_path,
        stdout_text="FINAL REPORT:\nDone",
        stderr_text="",
        output_text="FINAL REPORT:\nDone",
        error_text="",
        exit_code=0,
        timeout_seconds=300,
        stop_reason="process_exit",
        constraint_ledger=ledger,
    )

    assert evidence["constraint_compliance"]["status"] == "pass"
    assert evidence["evidence_result"]["status"] == "pass"


def test_constraint_compliance_allows_official_future_subject_when_sources_remain_in_window(tmp_path):
    output_dir = tmp_path / "output"
    output_dir.mkdir()
    (output_dir / "report.md").write_text(
        "Per the official roadmap summarized from in-window sources, the product plans a March 2027 launch.\n"
        "A July 2027 forecast appears in the source dated May 2026.\n"
        "Sources used were published through May 2026.\n"
    )
    ledger = build_constraint_ledger(
        instruction="Use sources from January 2024 through May 2026 only.",
        worker={"worker_id": "wrk_future_subject", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_future_subject",
    )

    evidence = build_run_evidence(
        worker={"worker_id": "wrk_future_subject", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_future_subject",
        runtime_name="codex-cli",
        model="gpt-test",
        command=["codex", "exec", "Analyze."],
        env={},
        workspace_dir=tmp_path,
        stdout_text="FINAL REPORT:\nDone",
        stderr_text="",
        output_text="Done",
        error_text="",
        exit_code=0,
        timeout_seconds=300,
        stop_reason="process_exit",
        constraint_ledger=ledger,
    )

    assert evidence["constraint_compliance"]["status"] == "pass"


def test_constraint_compliance_allows_explicitly_rejected_out_of_scope_sources(tmp_path):
    research_dir = tmp_path / "research"
    research_dir.mkdir()
    (research_dir / "rejected-sources.md").write_text(
        "Rejected out-of-scope evidence: Example source published June 2026; not used for facts, scoring, or deliverables.\n"
    )
    output_dir = tmp_path / "output"
    output_dir.mkdir()
    (output_dir / "report.md").write_text("FINAL REPORT:\nUsed only in-window evidence.\n")
    ledger = build_constraint_ledger(
        instruction="Use sources from January 2024 through May 2026 only.",
        worker={"worker_id": "wrk_rejected_source", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_rejected_source",
    )

    evidence = build_run_evidence(
        worker={"worker_id": "wrk_rejected_source", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_rejected_source",
        runtime_name="codex-cli",
        model="gpt-test",
        command=["codex", "exec", "Analyze."],
        env={},
        workspace_dir=tmp_path,
        stdout_text="FINAL REPORT:\nDone",
        stderr_text="",
        output_text="Done",
        error_text="",
        exit_code=0,
        timeout_seconds=300,
        stop_reason="process_exit",
        constraint_ledger=ledger,
    )

    assert evidence["constraint_compliance"]["status"] == "pass"


def test_constraint_compliance_still_flags_out_of_window_sources_only_marked_flagged(tmp_path):
    research_dir = tmp_path / "research"
    research_dir.mkdir()
    (research_dir / "notes.md").write_text("Flagged source published June 2026 appears in analysis notes.\n")
    ledger = build_constraint_ledger(
        instruction="Use sources from January 2024 through May 2026 only.",
        worker={"worker_id": "wrk_flagged_source", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_flagged_source",
    )

    evidence = build_run_evidence(
        worker={"worker_id": "wrk_flagged_source", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_flagged_source",
        runtime_name="codex-cli",
        model="gpt-test",
        command=["codex", "exec", "Analyze."],
        env={},
        workspace_dir=tmp_path,
        stdout_text="FINAL REPORT:\nDone",
        stderr_text="",
        output_text="Done",
        error_text="",
        exit_code=0,
        timeout_seconds=300,
        stop_reason="process_exit",
        constraint_ledger=ledger,
    )

    assert evidence["constraint_compliance"]["status"] == "fail"
    assert "June 2026" in evidence["constraint_compliance"]["issues"][0]["text"]


def test_constraint_compliance_skips_raw_support_snapshots_but_scans_research_notes(tmp_path):
    raw_support_dir = tmp_path / "research" / "site_snapshots" / "example.test"
    raw_support_dir.mkdir(parents=True)
    (raw_support_dir / "pages.json").write_text(
        json.dumps({"source": "Primary source published June 2026 and preserved only as a raw capture."})
    )
    ledger = build_constraint_ledger(
        instruction="Use sources from January 2024 through May 2026 only.",
        worker={"worker_id": "wrk_raw_support_source", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_raw_support_source",
    )

    clean_evidence = build_run_evidence(
        worker={"worker_id": "wrk_raw_support_source", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_raw_support_source",
        runtime_name="codex-cli",
        model="gpt-test",
        command=["codex", "exec", "Analyze."],
        env={},
        workspace_dir=tmp_path,
        stdout_text="FINAL REPORT:\nDone",
        stderr_text="",
        output_text="FINAL REPORT:\nDone",
        error_text="",
        exit_code=0,
        timeout_seconds=300,
        stop_reason="process_exit",
        constraint_ledger=ledger,
    )
    assert clean_evidence["constraint_compliance"]["status"] == "pass"

    (tmp_path / "research" / "notes.md").write_text("Primary source published June 2026 appears in notes.\n")
    dirty_evidence = build_run_evidence(
        worker={"worker_id": "wrk_raw_support_source", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_raw_support_source_dirty",
        runtime_name="codex-cli",
        model="gpt-test",
        command=["codex", "exec", "Analyze."],
        env={},
        workspace_dir=tmp_path,
        stdout_text="FINAL REPORT:\nDone",
        stderr_text="",
        output_text="FINAL REPORT:\nDone",
        error_text="",
        exit_code=0,
        timeout_seconds=300,
        stop_reason="process_exit",
        constraint_ledger=ledger,
    )
    assert dirty_evidence["constraint_compliance"]["status"] == "fail"
    assert dirty_evidence["constraint_compliance"]["issues"][0]["path"] == "research/notes.md"


def test_candidate_artifact_paths_sorts_before_max_entry_cap(tmp_path):
    for index in range(25):
        (tmp_path / f"root-{index:02d}.txt").write_text("lower priority")
    output_dir = tmp_path / "output"
    output_dir.mkdir()
    priority_artifact = output_dir / "final.csv"
    priority_artifact.write_text("name,value\nAlpha,1\n")

    paths = candidate_artifact_paths({"workspace_dir": str(tmp_path)}, max_entries=1)

    assert paths == [priority_artifact]


def test_run_evidence_records_html_browser_smoke_prerequisite(tmp_path, monkeypatch):
    output_dir = tmp_path / "output"
    output_dir.mkdir()
    (output_dir / "index.html").write_text(
        "<!doctype html><title>Smoke</title><main>Ready&nbsp;now</main><script>window.app=function(){return true}</script>"
    )
    monkeypatch.setattr("workers_projects_runtime.run_evidence.shutil.which", lambda _name: None)

    evidence = build_run_evidence(
        worker={"worker_id": "wrk_html", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_html",
        runtime_name="codex-cli",
        model="gpt-test",
        command=["codex", "exec", "Create HTML."],
        env={},
        workspace_dir=tmp_path,
        stdout_text="FINAL REPORT:\nCreated HTML",
        stderr_text="",
        output_text="Created HTML",
        error_text="",
        exit_code=0,
        timeout_seconds=None,
        stop_reason="process_exit",
        constraint_ledger=None,
    )

    html = next(item for item in evidence["artifacts"]["items"] if item["path"] == "output/index.html")
    assert html["html_validation"]["status"] == "unavailable"
    assert evidence["visual_render_evidence"]["status"] == "unavailable"
    assert evidence["content_hygiene"]["status"] == "pass"


def test_run_evidence_flags_notes_only_when_required_artifacts_are_missing(tmp_path):
    research_dir = tmp_path / "research"
    research_dir.mkdir()
    (research_dir / "batch-01.md").write_text("Research notes only.\nAlpha Capital appears here.\n")
    ledger = build_constraint_ledger(
        instruction="Include seed firms Alpha Capital and Beta Partners. Deliver PDF and XLSX artifacts.",
        worker={"worker_id": "wrk_notes_only", "profile": "claude-code", "execution_mode": "host"},
        run_id="run_notes_only",
    )

    evidence = build_run_evidence(
        worker={"worker_id": "wrk_notes_only", "profile": "claude-code", "execution_mode": "host"},
        run_id="run_notes_only",
        runtime_name="claude-code",
        model="claude-opus-4-8",
        command=["claude", "-p", "--effort", "max", "Analyze."],
        env={},
        workspace_dir=tmp_path,
        stdout_text="",
        stderr_text="Interrupted",
        output_text="",
        error_text="Interrupted",
        exit_code=None,
        timeout_seconds=300,
        stop_reason="timeout",
        constraint_ledger=ledger,
    )

    completion = evidence["completion_compliance"]
    assert completion["status"] == "fail"
    assert completion["notes_only"] is True
    assert set(completion["missing_required_artifact_types"]) == {"pdf", "xlsx"}
    assert completion["seed_entity_coverage"]["mentioned_count"] == 1
    assert "Beta Partners" in completion["seed_entity_coverage"]["missing"]
    assert evidence["evidence_result"]["status"] == "fail"


def test_run_evidence_flags_notes_only_when_deliverable_intent_has_no_extension(tmp_path):
    workspace = tmp_path / "workspace"
    research_dir = workspace / "research"
    research_dir.mkdir(parents=True)
    (research_dir / "batch-01.md").write_text("Research notes only; no user-facing report was produced.\n")
    ledger = build_constraint_ledger(
        instruction=(
            "### OUTPUT FORMAT\n"
            "Section 1 - Master Summary Table. All targets, sortable by weighted fit score.\n"
            "Section 2 - Deep Profiles. For each Tier 1 candidate, provide a narrative profile.\n"
            "Section 3 - Recommendation. Deliver the final report to the user."
        ),
        worker={"worker_id": "wrk_notes_only_intent", "profile": "claude-code", "execution_mode": "host"},
        run_id="run_notes_only_intent",
    )

    evidence = build_run_evidence(
        worker={"worker_id": "wrk_notes_only_intent", "profile": "claude-code", "execution_mode": "host"},
        run_id="run_notes_only_intent",
        runtime_name="claude-code",
        model="claude-test",
        command=["claude", "-p"],
        env={},
        workspace_dir=workspace,
        stdout_text='{"type":"result","result":"FINAL REPORT:\\nDone"}',
        stderr_text="",
        output_text="FINAL REPORT:\nDone",
        error_text="",
        exit_code=0,
        timeout_seconds=None,
        stop_reason="process_exit",
        constraint_ledger=ledger,
    )

    completion = evidence["completion_compliance"]
    assert completion["required_artifact_types"] == []
    assert completion["required_deliverable_intent"] is True
    assert completion["notes_only"] is True
    assert completion["status"] == "fail"
    assert any(issue["reason"] == "only support notes/planning artifacts were found" for issue in completion["issues"])
    assert evidence["evidence_result"]["status"] == "fail"


def test_run_evidence_allows_internal_notes_when_final_answer_requested_in_chat(tmp_path):
    workspace = tmp_path / "workspace"
    notes_dir = workspace / "notes"
    notes_dir.mkdir(parents=True)
    (notes_dir / "scratch.md").write_text("Internal notes used to prepare the answer.\n")
    ledger = build_constraint_ledger(
        instruction="Report back with the answer in chat. Do not create files.",
        worker={"worker_id": "wrk_chat_answer", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_chat_answer",
    )

    evidence = build_run_evidence(
        worker={"worker_id": "wrk_chat_answer", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_chat_answer",
        runtime_name="codex-cli",
        model="gpt-test",
        command=["codex", "exec"],
        env={},
        workspace_dir=workspace,
        stdout_text='{"type":"item.completed","item":{"type":"agent_message","text":"FINAL REPORT:\\nAnswer delivered inline."}}',
        stderr_text="",
        output_text="FINAL REPORT:\nAnswer delivered inline.",
        error_text="",
        exit_code=0,
        timeout_seconds=None,
        stop_reason="process_exit",
        constraint_ledger=ledger,
    )

    completion = evidence["completion_compliance"]
    assert completion["notes_only"] is True
    assert completion["required_deliverable_intent"] is False
    assert completion["status"] == "pass"
    assert evidence["evidence_result"]["status"] == "pass"


def test_run_evidence_completion_compliance_passes_for_requested_deliverables(tmp_path):
    output_dir = tmp_path / "output"
    output_dir.mkdir()
    (output_dir / "summary.csv").write_text("name,notes\nAlpha Capital,done\nBeta Partners,done\n")
    _write_minimal_pdf(output_dir / "report.pdf")
    openpyxl = pytest.importorskip("openpyxl")
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["name", "notes"])
    sheet.append(["Alpha Capital", "done"])
    sheet.append(["Beta Partners", "done"])
    workbook.save(output_dir / "screen.xlsx")
    ledger = build_constraint_ledger(
        instruction="Include seed firms Alpha Capital and Beta Partners. Deliver PDF, XLSX, and CSV artifacts.",
        worker={"worker_id": "wrk_complete", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_complete",
    )

    evidence = build_run_evidence(
        worker={"worker_id": "wrk_complete", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_complete",
        runtime_name="codex-cli",
        model="gpt-test",
        command=["codex", "exec", "-c", 'model_reasoning_effort="xhigh"', "Analyze."],
        env={},
        workspace_dir=tmp_path,
        stdout_text="FINAL REPORT:\nDone",
        stderr_text="",
        output_text="FINAL REPORT:\nDone",
        error_text="",
        exit_code=0,
        timeout_seconds=300,
        stop_reason="process_exit",
        constraint_ledger=ledger,
    )

    completion = evidence["completion_compliance"]
    assert completion["status"] == "pass"
    assert completion["missing_required_artifact_types"] == []
    assert completion["notes_only"] is False
    assert completion["seed_entity_coverage"]["missing"] == []
    assert evidence["evidence_result"]["status"] == "pass"


def test_run_evidence_hygiene_scans_csv_cells_not_whole_file(tmp_path):
    output_dir = tmp_path / "output"
    output_dir.mkdir()
    rows = ["id,entity,finding,rationale"]
    for index in range(1, 80):
        rows.append(f"{index},Alpha Robotics,Compact clean finding {index},Short rationale {index}")
    (output_dir / "screen.csv").write_text("\n".join(rows) + "\n")

    evidence = build_run_evidence(
        worker={"worker_id": "wrk_csv_cells", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_csv_cells",
        runtime_name="codex-cli",
        model="gpt-test",
        command=["codex", "exec", "Create CSV."],
        env={},
        workspace_dir=tmp_path,
        stdout_text="FINAL REPORT:\nDone",
        stderr_text="",
        output_text="FINAL REPORT:\nDone",
        error_text="",
        exit_code=0,
        timeout_seconds=300,
        stop_reason="process_exit",
        constraint_ledger=None,
    )

    assert evidence["content_hygiene"]["status"] == "pass"
    assert evidence["evidence_result"]["status"] == "pass"


def test_run_evidence_result_fails_invalid_professional_artifact(tmp_path):
    output_dir = tmp_path / "output"
    output_dir.mkdir()
    (output_dir / "report.pdf").write_text("not a pdf")

    evidence = build_run_evidence(
        worker={"worker_id": "wrk_invalid_artifact", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_invalid_artifact",
        runtime_name="codex-cli",
        model="gpt-test",
        command=["codex", "exec", "Create a PDF."],
        env={},
        workspace_dir=tmp_path,
        stdout_text="FINAL REPORT:\nDone",
        stderr_text="",
        output_text="FINAL REPORT:\nDone",
        error_text="",
        exit_code=0,
        timeout_seconds=300,
        stop_reason="process_exit",
        constraint_ledger=None,
    )

    summary = summarize_run_evidence_result(evidence)
    assert evidence["evidence_result"]["status"] == "fail"
    assert summary["status"] == "fail"
    assert any(item["reason"] == "professional artifact validation failed" for item in summary["failure_reasons"])


def test_constraint_ledger_does_not_treat_general_must_include_as_seed_entities():
    ledger = build_constraint_ledger(
        instruction=(
            "Create output/summary.csv and output/table.xlsx. "
            "The CSV and XLSX must include columns name, product, target_sector, score, note "
            "and two rows for Alpha Storage and Beta Grid."
        ),
        worker={"worker_id": "wrk_seed_false_positive", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_seed_false_positive",
    )

    assert ledger["outputs"]["format_expectations"] == ["xlsx", "csv"]
    assert ledger["seed_entities_or_files"] == []


def test_completion_compliance_does_not_require_attached_input_format_as_output(tmp_path):
    output_dir = tmp_path / "output"
    output_dir.mkdir()
    (output_dir / "answer.md").write_text("FINAL REPORT:\nCompared the attached spreadsheet and answered in chat.\n")
    ledger = build_constraint_ledger(
        instruction="Compare the attached CSV and answer in chat. Do not create extra files.",
        worker={"worker_id": "wrk_input_csv", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_input_csv",
    )

    evidence = build_run_evidence(
        worker={"worker_id": "wrk_input_csv", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_input_csv",
        runtime_name="codex-cli",
        model="gpt-test",
        command=["codex", "exec", "Compare."],
        env={},
        workspace_dir=tmp_path,
        stdout_text="FINAL REPORT:\nDone",
        stderr_text="",
        output_text="FINAL REPORT:\nDone",
        error_text="",
        exit_code=0,
        timeout_seconds=300,
        stop_reason="process_exit",
        constraint_ledger=ledger,
    )

    assert ledger["outputs"]["format_expectations"] == []
    assert "csv" not in evidence["completion_compliance"]["required_artifact_types"]
    assert evidence["completion_compliance"]["status"] == "pass"
    assert evidence["evidence_result"]["status"] == "pass"


def test_completion_compliance_does_not_require_uploads_path_input_format_as_output(tmp_path):
    artifacts = tmp_path / "artifacts"
    artifacts.mkdir()
    (artifacts / "glasshive-upload-smoke-result.txt").write_text(
        "FINAL REPORT:\nPDF bytes verified; HTML report created.\n",
        encoding="utf-8",
    )
    (artifacts / "glasshive-upload-smoke-report.html").write_text(
        "<!doctype html><title>Upload smoke</title><p>PDF bytes verified.</p>\n",
        encoding="utf-8",
    )
    ledger = build_constraint_ledger(
        instruction=(
            "Read uploads/glasshive-live-upload-smoke.pdf from the workspace. "
            "Create artifacts/glasshive-upload-smoke-result.txt and "
            "artifacts/glasshive-upload-smoke-report.html."
        ),
        worker={"worker_id": "wrk_upload_path_pdf", "profile": "codex-cli", "execution_mode": "docker"},
        run_id="run_upload_path_pdf",
    )

    evidence = build_run_evidence(
        worker={"worker_id": "wrk_upload_path_pdf", "profile": "codex-cli", "execution_mode": "docker"},
        run_id="run_upload_path_pdf",
        runtime_name="codex-cli",
        model="gpt-test",
        command=["codex", "exec", "Verify uploaded PDF bytes."],
        env={},
        workspace_dir=tmp_path,
        stdout_text="FINAL REPORT:\nDone",
        stderr_text="",
        output_text="FINAL REPORT:\nDone",
        error_text="",
        exit_code=0,
        timeout_seconds=300,
        stop_reason="process_exit",
        constraint_ledger=ledger,
    )

    assert set(ledger["outputs"]["format_expectations"]) == {"txt", "html"}
    assert "pdf" not in evidence["completion_compliance"]["required_artifact_types"]
    assert evidence["completion_compliance"]["missing_required_artifact_types"] == []
    assert evidence["completion_compliance"]["status"] == "pass"


def test_completion_compliance_accepts_context_named_input_pdf_with_txt_html_outputs(tmp_path):
    uploads = tmp_path / "uploads"
    uploads.mkdir()
    (uploads / "glasshive-ui-upload-smoke.pdf").write_bytes(
        b"%PDF-1.4\nGLASSHIVE_UPLOAD_STORAGE_OWNER_SMOKE_20260624\n"
    )
    (tmp_path / "upload-inspection-report.txt").write_text(
        "FINAL REPORT:\n"
        "Filename: glasshive-ui-upload-smoke.pdf\n"
        "Byte count: 58\n"
        "Starts with %PDF: yes\n"
        "Marker: GLASSHIVE_UPLOAD_STORAGE_OWNER_SMOKE_20260624\n",
        encoding="utf-8",
    )
    (tmp_path / "upload-inspection-report.html").write_text(
        "<!doctype html><title>Upload inspection</title>"
        "<p>glasshive-ui-upload-smoke.pdf starts with %PDF and contains "
        "GLASSHIVE_UPLOAD_STORAGE_OWNER_SMOKE_20260624.</p>\n",
        encoding="utf-8",
    )
    worker = {
        "worker_id": "wrk_context_named_input_pdf",
        "profile": "codex-cli",
        "execution_mode": "docker",
        "bootstrap_bundle_json": json.dumps(
            {"files": [{"path": "uploads/glasshive-ui-upload-smoke.pdf"}]}
        ),
    }
    ledger = build_constraint_ledger(
        instruction=(
            "Project description:\n"
            "Inspect the attached PDF file inside the GlassHive workspace and confirm whether the uploaded file is available there. "
            "Create a small TXT artifact and HTML artifact that report: filename, byte count, sha256, whether it starts with %PDF, "
            "and whether it contains marker GLASSHIVE_UPLOAD_STORAGE_OWNER_SMOKE_20260624. Return the results as generated artifacts in the workspace.\n\n"
            "Context:\n"
            "The attached uploaded file is named glasshive-ui-upload-smoke.pdf. The worker must verify that the uploaded file is actually available inside the GlassHive workspace, "
            "ideally by locating and reading the materialized uploaded file under uploads/glasshive-ui-upload-smoke.pdf or the host-provided upload location available in the workspace. "
            "Inspect the real file bytes rather than relying only on extracted text. Report in both a TXT artifact and an HTML artifact: filename, byte count, sha256, "
            "whether the file starts with the bytes/string %PDF, and whether it contains the marker string GLASSHIVE_UPLOAD_STORAGE_OWNER_SMOKE_20260624. "
            "Also include confirmation in the output about whether the uploaded file was available inside the GlassHive workspace. "
            "After generating the TXT and HTML files, keep them in the workspace as artifacts for download."
        ),
        worker=worker,
        run_id="run_context_named_input_pdf",
    )

    evidence = build_run_evidence(
        worker=worker,
        run_id="run_context_named_input_pdf",
        runtime_name="codex-cli",
        model="gpt-test",
        command=["codex", "exec", "Inspect upload."],
        env={},
        workspace_dir=tmp_path,
        stdout_text="FINAL REPORT:\nCreated upload-inspection-report.txt and upload-inspection-report.html",
        stderr_text="",
        output_text="FINAL REPORT:\nCreated upload-inspection-report.txt and upload-inspection-report.html",
        error_text="",
        exit_code=0,
        timeout_seconds=300,
        stop_reason="process_exit",
        constraint_ledger=ledger,
    )

    assert set(ledger["outputs"]["format_expectations"]) == {"txt", "html"}
    assert "pdf" not in evidence["completion_compliance"]["required_artifact_types"]
    assert evidence["completion_compliance"]["missing_required_artifact_types"] == []
    assert "confirmation in the" not in evidence["completion_compliance"]["seed_entity_coverage"].get("missing", [])
    assert evidence["completion_compliance"]["status"] == "pass"
    assert evidence["evidence_result"]["status"] == "pass"


def test_completion_compliance_subtracts_explicitly_forbidden_output_format(tmp_path):
    output_dir = tmp_path / "output"
    output_dir.mkdir()
    (output_dir / "answer.md").write_text("FINAL REPORT:\nInline answer only.\n")
    ledger = build_constraint_ledger(
        instruction="Do not create a PDF; answer inline as Markdown text.",
        worker={"worker_id": "wrk_no_pdf", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_no_pdf",
    )

    evidence = build_run_evidence(
        worker={"worker_id": "wrk_no_pdf", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_no_pdf",
        runtime_name="codex-cli",
        model="gpt-test",
        command=["codex", "exec", "Answer inline."],
        env={},
        workspace_dir=tmp_path,
        stdout_text="FINAL REPORT:\nDone",
        stderr_text="",
        output_text="FINAL REPORT:\nDone",
        error_text="",
        exit_code=0,
        timeout_seconds=300,
        stop_reason="process_exit",
        constraint_ledger=ledger,
    )

    assert ledger["outputs"]["format_expectations"] == []
    assert "pdf" in ledger["outputs"]["forbidden_format_expectations"]
    assert "pdf" not in evidence["completion_compliance"]["required_artifact_types"]
    assert evidence["completion_compliance"]["status"] == "pass"
    assert evidence["evidence_result"]["status"] == "pass"


def test_completion_compliance_does_not_count_zero_byte_required_artifact(tmp_path):
    output_dir = tmp_path / "output"
    output_dir.mkdir()
    (output_dir / "results.csv").write_text("")
    ledger = build_constraint_ledger(
        instruction="Deliver a CSV report.",
        worker={"worker_id": "wrk_empty_csv", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_empty_csv",
    )

    evidence = build_run_evidence(
        worker={"worker_id": "wrk_empty_csv", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_empty_csv",
        runtime_name="codex-cli",
        model="gpt-test",
        command=["codex", "exec", "Create CSV."],
        env={},
        workspace_dir=tmp_path,
        stdout_text="FINAL REPORT:\nDone",
        stderr_text="",
        output_text="FINAL REPORT:\nDone",
        error_text="",
        exit_code=0,
        timeout_seconds=300,
        stop_reason="process_exit",
        constraint_ledger=ledger,
    )

    assert evidence["artifacts"]["items"] == []
    assert evidence["completion_compliance"]["missing_required_artifact_types"] == ["csv"]
    assert evidence["completion_compliance"]["status"] == "fail"
    assert evidence["evidence_result"]["status"] == "fail"


def test_completion_compliance_requires_output_format_but_not_uploaded_input_format(tmp_path):
    output_dir = tmp_path / "output"
    output_dir.mkdir()
    (output_dir / "brief.pdf").write_bytes(b"%PDF-1.7\n% synthetic\n")
    ledger = build_constraint_ledger(
        instruction="Create a PDF report from the uploaded CSV.",
        worker={"worker_id": "wrk_pdf_from_csv", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_pdf_from_csv",
    )

    evidence = build_run_evidence(
        worker={"worker_id": "wrk_pdf_from_csv", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_pdf_from_csv",
        runtime_name="codex-cli",
        model="gpt-test",
        command=["codex", "exec", "Create report."],
        env={},
        workspace_dir=tmp_path,
        stdout_text="FINAL REPORT:\nDone",
        stderr_text="",
        output_text="FINAL REPORT:\nDone",
        error_text="",
        exit_code=0,
        timeout_seconds=300,
        stop_reason="process_exit",
        constraint_ledger=ledger,
    )

    assert evidence["completion_compliance"]["required_artifact_types"] == ["pdf"]
    assert "csv" not in evidence["completion_compliance"]["required_artifact_types"]
    assert evidence["completion_compliance"]["missing_required_artifact_types"] == []


def test_completion_compliance_ignores_seed_section_heading(tmp_path):
    workspace = tmp_path / "workspace"
    output = workspace / "output"
    output.mkdir(parents=True)
    (output / "report.md").write_text("Alpha Capital and Beta Partners are covered.\n")
    ledger = build_constraint_ledger(
        instruction="### SEED FIRM LIST\n(Expand beyond this list.)\nAlpha Capital, Beta Partners\nDeliver a report.",
        worker={"worker_id": "wrk_seed_heading", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_seed_heading",
    )

    evidence = build_run_evidence(
        worker={"worker_id": "wrk_seed_heading", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_seed_heading",
        runtime_name="codex-cli",
        model="gpt-test",
        command=["codex", "exec"],
        env={},
        workspace_dir=workspace,
        stdout_text='{"type":"item.completed","item":{"type":"agent_message","text":"FINAL REPORT:\\nDone"}}',
        stderr_text="",
        output_text="FINAL REPORT:\nDone",
        error_text="",
        exit_code=0,
        timeout_seconds=None,
        stop_reason="process_exit",
        constraint_ledger=ledger,
    )

    completion = evidence["completion_compliance"]
    assert completion["seed_entity_coverage"]["missing"] == []
    assert not any("FIRM LIST" in str(issue) for issue in completion["issues"])


def test_constraint_ledger_extracts_seed_block_terms_without_heading_noise():
    ledger = build_constraint_ledger(
        instruction=(
            "### SEED TARGET LIST\n"
            "(Expand beyond this list - do not limit to it. Aim for 50-75 total before scoring.)\n\n"
            "**Large Targets**\n"
            "Alpha Capital, Beta Partners, Gamma Growth\n"
            "**Focused Targets**\n"
            "Delta Health; Epsilon Software and Zeta Services\n"
        ),
        worker={"worker_id": "wrk_seed_block", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_seed_block",
    )

    seeds = ledger["seed_entities_or_files"]
    assert "Alpha Capital, Beta Partners, Gamma Growth" in seeds
    assert "Delta Health; Epsilon Software and Zeta Services" in seeds
    assert not any("SEED TARGET LIST" in seed for seed in seeds)
    assert not any("Aim for" in seed for seed in seeds)
    assert ledger["coverage_expectations"][0]["minimum"] == 50
    assert ledger["coverage_expectations"][0]["maximum"] == 75


def test_completion_compliance_handles_seed_topic_descriptor(tmp_path):
    workspace = tmp_path / "workspace"
    output = workspace / "output"
    output.mkdir(parents=True)
    (output / "brief.md").write_text(
        "Container sandbox controls and local MCP transport auth are both covered.\n"
    )
    ledger = build_constraint_ledger(
        instruction=(
            "Seed topics: container sandbox controls; local MCP transport auth. "
            "Deliver Markdown brief."
        ),
        worker={"worker_id": "wrk_seed_topic", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_seed_topic",
    )

    evidence = build_run_evidence(
        worker={"worker_id": "wrk_seed_topic", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_seed_topic",
        runtime_name="codex-cli",
        model="gpt-test",
        command=["codex", "exec"],
        env={},
        workspace_dir=workspace,
        stdout_text='{"type":"item.completed","item":{"type":"agent_message","text":"FINAL REPORT:\\nDone"}}',
        stderr_text="",
        output_text="FINAL REPORT:\nDone",
        error_text="",
        exit_code=0,
        timeout_seconds=None,
        stop_reason="process_exit",
        constraint_ledger=ledger,
    )

    completion = evidence["completion_compliance"]
    assert completion["seed_entity_coverage"]["missing"] == []
    assert completion["status"] == "pass"
    assert evidence["evidence_result"]["status"] == "pass"


def test_coverage_compliance_fails_when_table_rows_below_requested_range(tmp_path):
    workspace = tmp_path / "workspace"
    output = workspace / "output"
    output.mkdir(parents=True)
    with (output / "screen.csv").open("w", encoding="utf-8") as handle:
        handle.write("entity,score\n")
        for index in range(36):
            handle.write(f"Entity {index},4.{index % 10}\n")
    ledger = build_constraint_ledger(
        instruction="Expand beyond the seed list. Aim for 50-75 firms total before scoring. Deliver CSV screen.",
        worker={"worker_id": "wrk_coverage_low", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_coverage_low",
    )

    evidence = build_run_evidence(
        worker={"worker_id": "wrk_coverage_low", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_coverage_low",
        runtime_name="codex-cli",
        model="gpt-test",
        command=["codex", "exec"],
        env={},
        workspace_dir=workspace,
        stdout_text='{"type":"item.completed","item":{"type":"agent_message","text":"FINAL REPORT:\\nDone"}}',
        stderr_text="",
        output_text="FINAL REPORT:\nDone",
        error_text="",
        exit_code=0,
        timeout_seconds=None,
        stop_reason="process_exit",
        constraint_ledger=ledger,
    )

    coverage = evidence["coverage_compliance"]
    assert coverage["status"] == "fail"
    assert coverage["observed_max_count"] == 36
    assert any(issue["reason"] == "coverage count below requested minimum" for issue in coverage["issues"])
    assert evidence["evidence_result"]["status"] == "fail"


def test_coverage_compliance_passes_when_table_rows_match_requested_range(tmp_path):
    workspace = tmp_path / "workspace"
    output = workspace / "output"
    output.mkdir(parents=True)
    _write_xlsx_test_path = output / "screen.xlsx"
    try:
        import openpyxl  # type: ignore
    except Exception:
        pytest.skip("openpyxl unavailable")
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Master"
    sheet.append(["entity", "score"])
    for index in range(63):
        sheet.append([f"Entity {index}", 4.0])
    workbook.save(_write_xlsx_test_path)
    ledger = build_constraint_ledger(
        instruction="Expand beyond the seed list. Aim for 50-75 firms total before scoring. Deliver XLSX workbook.",
        worker={"worker_id": "wrk_coverage_pass", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_coverage_pass",
    )

    evidence = build_run_evidence(
        worker={"worker_id": "wrk_coverage_pass", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_coverage_pass",
        runtime_name="codex-cli",
        model="gpt-test",
        command=["codex", "exec"],
        env={},
        workspace_dir=workspace,
        stdout_text='{"type":"item.completed","item":{"type":"agent_message","text":"FINAL REPORT:\\nDone"}}',
        stderr_text="",
        output_text="FINAL REPORT:\nDone",
        error_text="",
        exit_code=0,
        timeout_seconds=None,
        stop_reason="process_exit",
        constraint_ledger=ledger,
    )

    coverage = evidence["coverage_compliance"]
    assert coverage["status"] == "pass"
    assert coverage["observed_max_count"] == 63
    assert evidence["evidence_result"]["status"] == "pass"


def test_coverage_compliance_passes_when_any_deliverable_table_matches_requested_range(tmp_path):
    workspace = tmp_path / "workspace"
    output = workspace / "output"
    output.mkdir(parents=True)
    with (output / "screen.csv").open("w", encoding="utf-8") as handle:
        handle.write("entity,score\n")
        for index in range(50):
            handle.write(f"Entity {index},4.{index % 10}\n")
    html_rows = "\n".join(f"<tr><td>Other {index}</td></tr>" for index in range(80))
    (output / "appendix.html").write_text(f"<table>{html_rows}</table>\n", encoding="utf-8")
    ledger = build_constraint_ledger(
        instruction="Expand beyond the seed list. Aim for 50-75 firms total before scoring. Deliver CSV screen.",
        worker={"worker_id": "wrk_coverage_any_match", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_coverage_any_match",
    )

    evidence = build_run_evidence(
        worker={"worker_id": "wrk_coverage_any_match", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_coverage_any_match",
        runtime_name="codex-cli",
        model="gpt-test",
        command=["codex", "exec"],
        env={},
        workspace_dir=workspace,
        stdout_text='{"type":"item.completed","item":{"type":"agent_message","text":"FINAL REPORT:\\nDone"}}',
        stderr_text="",
        output_text="FINAL REPORT:\nDone",
        error_text="",
        exit_code=0,
        timeout_seconds=None,
        stop_reason="process_exit",
        constraint_ledger=ledger,
    )

    coverage = evidence["coverage_compliance"]
    assert coverage["status"] == "pass"
    assert coverage["observed_max_count"] > 75
    assert coverage["matched_count"] == 50
    assert evidence["evidence_result"]["status"] == "pass"


def test_coverage_compliance_counts_structured_final_answer_items(tmp_path):
    ledger = build_constraint_ledger(
        instruction="Produce at least 3 items in the final answer. Answer in chat.",
        worker={"worker_id": "wrk_final_items", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_final_items",
    )
    output_text = "FINAL REPORT:\n1. Alpha\n2. Beta\n3. Gamma\n"

    evidence = build_run_evidence(
        worker={"worker_id": "wrk_final_items", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_final_items",
        runtime_name="codex-cli",
        model="gpt-test",
        command=["codex", "exec"],
        env={},
        workspace_dir=tmp_path,
        stdout_text='{"type":"item.completed","item":{"type":"agent_message","text":"FINAL REPORT:\\n1. Alpha\\n2. Beta\\n3. Gamma"}}',
        stderr_text="",
        output_text=output_text,
        error_text="",
        exit_code=0,
        timeout_seconds=None,
        stop_reason="process_exit",
        constraint_ledger=ledger,
    )

    coverage = evidence["coverage_compliance"]
    assert coverage["status"] == "pass"
    assert coverage["matched_count"] == 3
    assert any(item["kind"] == "final_output_list_items" for item in coverage["counts"])
    assert evidence["completion_compliance"]["required_deliverable_intent"] is False
    assert evidence["evidence_result"]["status"] == "pass"


def test_coverage_compliance_warns_when_count_unverifiable_in_professional_artifact(tmp_path):
    workspace = tmp_path / "workspace"
    output = workspace / "output"
    output.mkdir(parents=True)
    _write_minimal_pdf(output / "executive_briefing.pdf")
    ledger = build_constraint_ledger(
        instruction="Research 40-50 firms and deliver a PDF executive briefing.",
        worker={"worker_id": "wrk_coverage_pdf", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_coverage_pdf",
    )

    evidence = build_run_evidence(
        worker={"worker_id": "wrk_coverage_pdf", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_coverage_pdf",
        runtime_name="codex-cli",
        model="gpt-test",
        command=["codex", "exec"],
        env={},
        workspace_dir=workspace,
        stdout_text='{"type":"item.completed","item":{"type":"agent_message","text":"FINAL REPORT:\\nResearched 45 firms and created the PDF."}}',
        stderr_text="",
        output_text="FINAL REPORT:\nResearched 45 firms and created the PDF.",
        error_text="",
        exit_code=0,
        timeout_seconds=None,
        stop_reason="process_exit",
        constraint_ledger=ledger,
    )

    coverage = evidence["coverage_compliance"]
    assert coverage["status"] == "warn"
    assert any(item["reason"] == "coverage count could not be verified" for item in coverage["issues"])
    assert evidence["completion_compliance"]["status"] == "pass"
    assert evidence["evidence_result"]["status"] == "warn"


def test_completion_compliance_does_not_treat_seed_reference_sentence_as_seed_path(tmp_path):
    workspace = tmp_path / "workspace"
    output = workspace / "output"
    output.mkdir(parents=True)
    (output / "report.md").write_text("Alpha Robotics and Beta Energy are covered.\n")
    (output / "screen.csv").write_text("entity\nAlpha Robotics\nBeta Energy\n")
    ledger = build_constraint_ledger(
        instruction=(
            "Create output/report.md and output/screen.csv. "
            "The report and CSV must mention both seed entities. Do not create a PDF."
        ),
        worker={"worker_id": "wrk_seed_reference", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_seed_reference",
    )

    evidence = build_run_evidence(
        worker={"worker_id": "wrk_seed_reference", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_seed_reference",
        runtime_name="codex-cli",
        model="gpt-test",
        command=["codex", "exec"],
        env={},
        workspace_dir=workspace,
        stdout_text='{"type":"item.completed","item":{"type":"agent_message","text":"FINAL REPORT:\\nDone"}}',
        stderr_text="",
        output_text="FINAL REPORT:\nDone",
        error_text="",
        exit_code=0,
        timeout_seconds=None,
        stop_reason="process_exit",
        constraint_ledger=ledger,
    )

    coverage = evidence["completion_compliance"]["seed_entity_coverage"]
    assert coverage["status"] == "not_applicable"
    assert evidence["completion_compliance"]["missing_required_artifact_types"] == []
    assert evidence["evidence_result"]["status"] == "pass"


def test_completion_compliance_keeps_required_formats_when_line_also_forbids_pdf(tmp_path):
    workspace = tmp_path / "workspace"
    output = workspace / "output"
    output.mkdir(parents=True)
    (output / "report.md").write_text("Done.\n")
    (output / "screen.csv").write_text("entity\nAlpha Robotics\n")
    ledger = build_constraint_ledger(
        instruction="Deliver output/report.md and output/screen.csv. Do not create a PDF.",
        worker={"worker_id": "wrk_mixed_formats", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_mixed_formats",
    )

    evidence = build_run_evidence(
        worker={"worker_id": "wrk_mixed_formats", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_mixed_formats",
        runtime_name="codex-cli",
        model="gpt-test",
        command=["codex", "exec"],
        env={},
        workspace_dir=workspace,
        stdout_text='{"type":"item.completed","item":{"type":"agent_message","text":"FINAL REPORT:\\nDone"}}',
        stderr_text="",
        output_text="FINAL REPORT:\nDone",
        error_text="",
        exit_code=0,
        timeout_seconds=None,
        stop_reason="process_exit",
        constraint_ledger=ledger,
    )

    completion = evidence["completion_compliance"]
    assert completion["required_artifact_types"] == ["csv", "md"]
    assert completion["missing_required_artifact_types"] == []
    assert "pdf" not in completion["required_artifact_types"]
    assert evidence["evidence_result"]["status"] == "pass"


def test_completion_compliance_keeps_required_format_after_parenthetical_no_pdf(tmp_path):
    workspace = tmp_path / "workspace"
    output = workspace / "output"
    output.mkdir(parents=True)
    (output / "screen.csv").write_text("entity\nAlpha Robotics\n")
    ledger = build_constraint_ledger(
        instruction="Deliver a report (no PDF) in CSV.",
        worker={"worker_id": "wrk_parenthetical_formats", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_parenthetical_formats",
    )

    evidence = build_run_evidence(
        worker={"worker_id": "wrk_parenthetical_formats", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_parenthetical_formats",
        runtime_name="codex-cli",
        model="gpt-test",
        command=["codex", "exec"],
        env={},
        workspace_dir=workspace,
        stdout_text='{"type":"item.completed","item":{"type":"agent_message","text":"FINAL REPORT:\\nDone"}}',
        stderr_text="",
        output_text="FINAL REPORT:\nDone",
        error_text="",
        exit_code=0,
        timeout_seconds=None,
        stop_reason="process_exit",
        constraint_ledger=ledger,
    )

    completion = evidence["completion_compliance"]
    assert completion["required_artifact_types"] == ["csv"]
    assert completion["missing_required_artifact_types"] == []
    assert "pdf" not in completion["required_artifact_types"]
    assert evidence["evidence_result"]["status"] == "pass"


def test_constraint_softening_only_scans_planning_files(tmp_path):
    workspace = tmp_path / "workspace"
    output = workspace / "output"
    notes = workspace / "notes"
    output.mkdir(parents=True)
    notes.mkdir(parents=True)
    (output / "report.md").write_text("This is not a broad transformation program; it is a focused result.\n")
    (notes / "plan.md").write_text("Use sources through May 2026 wherever possible.\n")
    ledger = build_constraint_ledger(
        instruction="Use sources through May 2026 only.",
        worker={"worker_id": "wrk_softening", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_softening",
    )

    evidence = build_run_evidence(
        worker={"worker_id": "wrk_softening", "profile": "codex-cli", "execution_mode": "host"},
        run_id="run_softening",
        runtime_name="codex-cli",
        model="gpt-test",
        command=["codex", "exec"],
        env={},
        workspace_dir=workspace,
        stdout_text='{"type":"item.completed","item":{"type":"agent_message","text":"FINAL REPORT:\\nDone"}}',
        stderr_text="",
        output_text="FINAL REPORT:\nDone",
        error_text="",
        exit_code=0,
        timeout_seconds=None,
        stop_reason="process_exit",
        constraint_ledger=ledger,
    )

    issues = evidence["constraint_compliance"]["issues"]
    assert len(issues) == 1
    assert issues[0]["path"] == "notes/plan.md"


def test_text_artifact_payload_walks_json_leaves_for_hygiene(tmp_path):
    workspace = tmp_path / "workspace"
    research = workspace / "research"
    research.mkdir(parents=True)
    rows = [
        {
            "firm": f"Firm {index}",
            "notes": "normal sourced row",
            "structuralCase": "ordinary narrative profile prose " * 50,
        }
        for index in range(80)
    ]
    (research / "dataset.json").write_text(json.dumps(rows))

    payload = _text_artifact_payload(workspace)
    result = check_content_hygiene(payload)

    assert "research/dataset.json" not in payload
    assert result["status"] == "pass"


def test_text_artifact_payload_skips_raw_support_snapshots_but_scans_deliverables(tmp_path):
    workspace = tmp_path / "workspace"
    raw_support_dir = workspace / "research" / "site_snapshots" / "example.test"
    output_dir = workspace / "output"
    raw_support_dir.mkdir(parents=True)
    output_dir.mkdir(parents=True)
    (raw_support_dir / "pages.json").write_text(
        json.dumps({"page": "Skip to main content. Privacy Policy. Please enable JavaScript."})
    )
    (output_dir / "findings.csv").write_text("entity,notes\nAlpha,Privacy Policy copied into the final cell\n")

    payload = _text_artifact_payload(workspace)
    result = check_content_hygiene(payload)

    assert all("research/site_snapshots" not in path for path in payload)
    assert any(path.startswith("output/findings.csv") for path in payload)
    assert result["status"] == "warn"
    assert result["issues"][0]["path"].startswith("output/findings.csv")
